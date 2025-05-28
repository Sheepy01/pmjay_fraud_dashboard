from .models import Last24Hour, SuspiciousHospital, HospitalBeds, UploadHistory
from openpyxl.styles import PatternFill, Font, Border, Side
from django.contrib.auth.decorators import login_required
from django.db.models import Case, When, Value, CharField
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_POST
from django.contrib.auth import authenticate, login, logout
from django.views.decorators.http import require_http_methods
from django.template.loader import render_to_string
from django.db.models import Sum, Q, F, Func, Count, Subquery, Max, OuterRef, Value, BooleanField, Exists, IntegerField
from datetime import date
import random
from django.http import JsonResponse, HttpResponse
import re
from django.utils.timezone import timedelta
from datetime import date
from django.shortcuts import render, redirect
import datetime
from django.db.models.functions import Cast
from collections import defaultdict
from django.core.paginator import Paginator
import pandas as pd
from weasyprint import HTML
from datetime import timedelta
from django.shortcuts import render
from django.db.models.functions import TruncDate
from django.utils import timezone
import io
from django.contrib import messages
from django.db import transaction
import numpy as np
import sys

def login_view(request):
    # If they’re already logged in, send them straight to dashboard
    if request.user.is_authenticated:
        return redirect('dashboard')

    error = None
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=request.POST['username'], password=request.POST['password'])
        if user is not None:
            login(request, user)
            # call_command('process_new_files')
            return redirect('dashboard')
        else:
            error = "Invalid username or password"

    return render(request, 'login.html', {'error': error})

def logout_view(request):
    logout(request)
    return redirect('login')

class Upper(Func):
    function = 'UPPER'
    template = '%(function)s(%(expressions)s)'

@login_required
def import_data_view(request):
    if request.method == 'POST':
        uploaded_files = request.FILES.getlist('files')
        required_columns = [
            'Registration Id', 
            'Preauth Initiated Date',
            'Preauth Initiated Time',
            'Hospital Code',
            'Claim Initiated Amount(Rs.)',
            'Hospital Type',
            'Case Type',
            'State Name',
            'Age(Years)',
            'Procedure Code',
            'District Name',
            'Patient Name',
            'Gender',
            'Hospital Name',
            'Hospital State Name',
            'Family Id'
        ]

        try:
            new_records = 0
            updated_records = 0
            
            for uploaded_file in uploaded_files:
                try:
                    df = pd.read_excel(
                        uploaded_file,
                        sheet_name='Dump',
                        engine='openpyxl',
                        dtype={'Registration Id': str}
                    )
                    
                    # Validate and clean data
                    df.columns = df.columns.str.strip()
                    missing_cols = [col for col in required_columns if col not in df.columns]
                    if missing_cols:
                        messages.error(request, f"Skipped {uploaded_file.name}: Missing columns {', '.join(missing_cols)}")
                        continue

                    # Convert date and time
                    df['preauth_date'] = pd.to_datetime(
                        df['Preauth Initiated Date'],
                        dayfirst=True,  # Prioritize DD-MM-YYYY format
                        errors='coerce'
                    ).dt.date
                    
                    df['preauth_time'] = pd.to_datetime(
                        df['Preauth Initiated Time'].astype(str).str.replace(r'\s+[AP]M$', '', regex=True),
                        format='%H:%M:%S',  # Try with seconds first
                        errors='coerce'
                    ).dt.time

                    # Fallback for times without seconds
                    if df['preauth_time'].isna().any():
                        df['preauth_time'] = pd.to_datetime(
                            df['Preauth Initiated Time'].astype(str).str.replace(r'\s+[AP]M$', '', regex=True),
                            format='%H:%M',
                            errors='coerce'
                        ).dt.time
                    
                    df['Registration Id'] = (
                        df['Registration Id']
                        .astype(str)
                        .str.strip()
                        .replace({'nan': None, 'None': None, '': None})
                    )

                    # Show problematic rows
                    invalid_rows = df[
                        df['preauth_date'].isna() |
                        df['preauth_time'].isna() |
                        df['Registration Id'].isna()
                    ]

                    # Filter valid records
                    valid_df = df.dropna(subset=[
                        'preauth_date', 
                        'preauth_time', 
                        'Registration Id'
                    ])
                    
                    # Create model instances
                    instances = []
                    for _, row in valid_df.iterrows():
                        instances.append(Last24Hour(
                            registration_id=str(row['Registration Id']).strip(),
                            preauth_initiated_date=row['preauth_date'],
                            preauth_initiated_time=row['preauth_time'],
                            hospital_id=str(row['Hospital Code']).strip().upper(),  # Corrected
                            hospital_type=str(row['Hospital Type']).strip(),
                            case_type=str(row['Case Type']).strip(),
                            claim_initiated_amount=row['Claim Initiated Amount(Rs.)'],
                            state_name=str(row['State Name']).strip(),
                            age_years=str(row['Age(Years)']).strip(),
                            procedure_code=str(row['Procedure Code']).strip(),
                            district_name=str(row['District Name']).strip(),
                            patient_name=str(row['Patient Name']).strip(),
                            gender=str(row['Gender']).strip().upper(),
                            hospital_name=str(row['Hospital Name']).strip().upper(),
                            hospital_state_name=str(row['Hospital State Name']).strip().upper(),
                            family_id=str(row['Family Id']).strip(),
                        ))

                    # Bulk create with conflict handling
                    result = Last24Hour.objects.bulk_create(
                        instances,
                        update_conflicts=True,
                        unique_fields=['registration_id', 'preauth_initiated_date', 'preauth_initiated_time'],
                        update_fields=[
                            'hospital_id',  # Not hospital_code
                            'hospital_type',
                            'case_type',
                            'claim_initiated_amount',
                            'state_name',
                            'age_years',
                            'procedure_code',
                            'district_name',
                            'patient_name',
                            'gender',
                            'hospital_name',
                            'hospital_state_name',
                            'family_id'
                        ])
                    
                    new_records += len(result)
                    updated_records += len(instances) - len(result)

                except Exception as e:
                    messages.error(request, f"Error processing {uploaded_file.name}: {str(e)}")
                    continue

            messages.success(request, 
                f"Processed {new_records} new and {updated_records} updated records"
            )
            
        except Exception as e:
            messages.error(request, f"System error: {str(e)}")

        return redirect('dashboard')
    
def data_management(request):
    histories = UploadHistory.objects.in_bulk(field_name='model_type')
    return render(request, 'data_management.html', {
        'active_page': 'data_management',
        'file_upload_histories': histories,
    })

@require_POST
def upload_management_data(request):
    file = request.FILES.get('file')
    model_type = request.POST.get('model_type')
    
    if not file:
        return JsonResponse({'status': 'error', 'message': 'No file uploaded'}, status=400)
    if not model_type:
        return JsonResponse({'status': 'error', 'message': 'Missing model type'}, status=400)

    try:
        df = pd.read_excel(file)
        df = df.replace({np.nan: None})

        required_columns = {
            'suspicious': ['Hospital Id', 'Hospital Name', 'Number of Surgeons', 'Number of OT'],
            'beds': ['Hospital ID', 'Bed Strength', 'Hospital Name']
        }.get(model_type, [])

        if not required_columns:
            return JsonResponse({'status': 'error', 'message': 'Invalid model type'}, status=400)

        missing_cols = [col for col in required_columns if col not in df.columns]
        if missing_cols:
            return JsonResponse({
                'status': 'error',
                'message': f'Missing columns: {", ".join(missing_cols)}'
            }, status=400)

        # Handle duplicates before processing
        if model_type == 'suspicious':
            # Remove duplicates keeping first occurrence
            initial_count = len(df)
            df = df.drop_duplicates(
                subset=['Hospital Id'], 
                keep='first' 
            )
            removed_duplicates = initial_count - len(df)
            
            with transaction.atomic():
                SuspiciousHospital.objects.all().delete()
                hospitals = [
                    SuspiciousHospital(
                        hospital_id=row['Hospital Id'],
                        hospital_name=row['Hospital Name'],
                        number_of_surgeons=row['Number of Surgeons'],
                        number_of_ot=row['Number of OT']
                    )
                    for _, row in df.iterrows()
                ]
                SuspiciousHospital.objects.bulk_create(hospitals)
                UploadHistory.objects.update_or_create(
                    model_type='suspicious',
                    defaults={'filename': file.name}
                )
                
                return JsonResponse({
                    'status': 'success',
                    'message': f'Uploaded {len(df)} records (removed {removed_duplicates} duplicates)'
                })

        elif model_type == 'beds':
            # Handle Hospital Beds upload
            required_columns = ['Hospital ID', 'Bed Strength']
            
            # Fill missing bed_strength with 0
            df['Bed Strength'] = df['Bed Strength'].fillna(0)
            
            # Convert to integers with error handling
            try:
                df['Bed Strength'] = df['Bed Strength'].astype(int)
            except ValueError:
                return JsonResponse({
                    'status': 'error',
                    'message': 'bed_strength contains non-numeric values that cannot be converted to integers'
                }, status=400)

            # Remove duplicates
            initial_count = len(df)
            df = df.drop_duplicates(subset=['Hospital ID'], keep='first')
            removed_duplicates = initial_count - len(df)

            with transaction.atomic():
                HospitalBeds.objects.all().delete()
                beds = []
                for index, row in df.iterrows():
                    try:
                        beds.append(HospitalBeds(
                            hospital_id=row['Hospital ID'],
                            hospital_name=row.get('Hospital Name', ''),
                            bed_strength=row['Bed Strength']
                        ))
                    except Exception as e:
                        return JsonResponse({
                            'status': 'error',
                            'message': f'Error in row {index+2}: {str(e)}'
                        }, status=400)
                        
                HospitalBeds.objects.bulk_create(beds)
                UploadHistory.objects.update_or_create(
                    model_type='beds',
                    defaults={'filename': file.name}
                )
                
            return JsonResponse({
                'status': 'success',
                'message': f'Uploaded {len(df)} beds records. {removed_duplicates} duplicates removed. {initial_count - len(df)} nulls converted to 0.'
            })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'{str(e)} (Line {sys.exc_info()[-1].tb_lineno if hasattr(sys, "exc_info") else "N/A"})'
        }, status=400)
    
@login_required
def latest_uploads(request):
    """
    Return JSON mapping of model_type -> {filename, uploaded_at}
    """
    qs = UploadHistory.objects.all()
    data = {}
    for hist in qs:
        data[hist.model_type] = {
            'filename': hist.filename,
            # ISO timestamp is easiest for JS to parse, but we can shorten if you like
            'uploaded_at': hist.uploaded_at.strftime('%Y-%m-%d %H:%M'),
        }
    return JsonResponse(data)

def get_management_data(request):
    model_type = request.GET.get('model_type')
    
    try:
        if model_type == 'suspicious':
            data = SuspiciousHospital.objects.all()
            fields = ['hospital_id', 'hospital_name', 'number_of_surgeons', 'number_of_ot']
            labels = ['Hospital ID', 'Hospital Name', 'Number of Surgeons', 'Number of OTs']
        elif model_type == 'beds':
            data = HospitalBeds.objects.all()
            fields = ['hospital_id', 'hospital_name', 'bed_strength']
            labels = ['Hospital ID', 'Hospital Name', 'Bed Strength']
        else:
            return HttpResponse('Invalid model type')
            
        # build the table head
        html = '<table><thead><tr>'
        for label in labels:
            html += f'<th>{label}</th>'
        html += '</tr></thead><tbody>'

        # build the rows
        for item in data:
            html += '<tr>'
            for field in fields:
                html += f'<td>{getattr(item, field)}</td>'
            html += '</tr>'
        html += '</tbody></table>'
    except Exception as e:
        return HttpResponse(f'Error: {str(e)}', status=400)
    
    return HttpResponse(html)

def get_districts(request):
    districts = Last24Hour.objects.values_list('district_name', flat=True).distinct()
    district_list = [d for d in districts if d]  # Remove None/empty values
    return JsonResponse({'districts': district_list})

def get_flagged_claims(request):
    # 1. parse district
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    # 2. parse dates
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    if startDate and endDate:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date()
        end_date   = datetime.datetime.strptime(endDate, '%Y-%m-%d').date()
    else:
        today = timezone.localdate()
        start_date = end_date = today

    # 3. build base queryset over the date range
    suspicious_hospitals = SuspiciousHospital.objects.values_list('hospital_id', flat=True)
    base_qs = Last24Hour.objects.filter(
        hospital_id__in=suspicious_hospitals,
        hospital_type='P',
        preauth_initiated_date__date__gte=start_date,
        preauth_initiated_date__date__lte=end_date
    )
    if districts:
        base_qs = base_qs.filter(district_name__in=districts)

    # 4. counts
    total = base_qs.count()
    # yesterday always relative to end_date?
    yesterday = end_date - timedelta(days=1)
    yesterday_count = base_qs.filter(preauth_initiated_date__date=yesterday).count()
    last_30_days = Last24Hour.objects.filter(
        hospital_id__in=suspicious_hospitals,
        hospital_type='P',
        preauth_initiated_date__date__gte=end_date - timedelta(days=30),
        preauth_initiated_date__date__lte=end_date
    )
    if districts:
        last_30_days = last_30_days.filter(district_name__in=districts)

    return JsonResponse({
        'total': total,
        'yesterday': yesterday_count,
        'last_30_days': last_30_days.count()
    })

def get_flagged_claims_details(request):
    district_param = request.GET.get('district', '')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 50))
    districts = district_param.split(',') if district_param else []

    # 1. Parse start_date / end_date from GET
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()

    # 2. Base queryset over the date range
    suspicious_hospitals = SuspiciousHospital.objects.values_list('hospital_id', flat=True)
    qs = Last24Hour.objects.filter(
        hospital_id__in=suspicious_hospitals,
        hospital_type='P',
        preauth_initiated_date__date__gte=start_date,
        preauth_initiated_date__date__lte=end_date
    )
    if districts:
        qs = qs.filter(district_name__in=districts)

    # 3. Pagination
    paginator = Paginator(qs.order_by('preauth_initiated_date'), page_size)
    page_obj = paginator.get_page(page)

    # 4. Build response data
    data = []
    for idx, case in enumerate(page_obj.object_list, 1):
        data.append({
            'serial_no': (page_obj.number - 1) * page_size + idx,
            'claim_id': case.registration_id or case.case_id or 'N/A',
            'patient_name': case.patient_name or f"Patient {case.member_id}",
            'district_name': case.district_name or 'N/A',
            'preauth_initiated_date': case.preauth_initiated_date.strftime('%Y-%m-%d') 
                                      if case.preauth_initiated_date else 'N/A',
            'preauth_initiated_time': case.preauth_initiated_time or 'N/A',
            'hospital_id': case.hospital_id or 'N/A',
            'hospital_name': case.hospital_name or 'N/A',
            'amount': float(case.claim_initiated_amount) if case.claim_initiated_amount else 0.0,
            'reason': 'Suspicious hospital'
        })

    return JsonResponse({
        'data': data,
        'pagination': {
            'total_records': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
        }
    })

# Add to views.py
def get_flagged_claims_by_district(request):
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()
    
    # Base queryset with today's filter
    queryset = Last24Hour.objects.filter(
        hospital_id__in=SuspiciousHospital.objects.values('hospital_id'),
        hospital_type='P',
        preauth_initiated_date__date__gte=start_date,
        preauth_initiated_date__date__lte=end_date  # Added today filter
    )
    
    if districts:
        queryset = queryset.filter(district_name__in=districts)
    
    # Aggregate data by district
    district_data = queryset.values('district_name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Prepare response data
    data = {
        'districts': [item['district_name'] or 'Unknown' for item in district_data],
        'counts': [item['count'] for item in district_data]
    }
    
    return JsonResponse(data)

def get_all_flagged_claims(request):
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []
    
    # Get current date filter
    today = date(2025, 2, 5)
    
    # Base query with today's filter
    suspicious_hospitals = SuspiciousHospital.objects.values_list('hospital_id', flat=True)
    flagged_cases = Last24Hour.objects.filter(
        Q(hospital_id__in=suspicious_hospitals) &
        Q(hospital_type='P') &
        Q(preauth_initiated_date__date=today)  # Added today filter
    )
    
    if districts:
        flagged_cases = flagged_cases.filter(district_name__in=districts)
    
    data = []
    for idx, case in enumerate(flagged_cases.order_by('preauth_initiated_date'), 1):
        data.append({
            'serial_no': idx,
            'claim_id': case.registration_id or case.case_id or 'N/A',
            'patient_name': case.patient_name or f"Patient {case.member_id}",
            'hospital_name': case.hospital_name or 'N/A',
            'district_name': case.district_name or 'N/A',
            'amount': float(case.claim_initiated_amount) if case.claim_initiated_amount else 0.0,
            'reason': 'Suspicious hospital'
        })
    
    return JsonResponse({'data': data})

def get_age_distribution(request):
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')

    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()

    queryset = Last24Hour.objects.filter(
        hospital_id__in=SuspiciousHospital.objects.values('hospital_id'),
        hospital_type='P',
        preauth_initiated_date__date__gte=start_date,  # Added today filter
        preauth_initiated_date__date__lte=end_date  # Added today filter
    )

    if districts:
        queryset = queryset.filter(district_name__in=districts)

    # Age groups aggregation
    age_groups = {
        '15-29': Count('id', filter=Q(age_years__gte=15, age_years__lte=29)),
        '30-44': Count('id', filter=Q(age_years__gte=30, age_years__lte=44)),
        '45-59': Count('id', filter=Q(age_years__gte=45, age_years__lte=59)),
        '60+': Count('id', filter=Q(age_years__gte=60))
    }

    age_data = queryset.aggregate(**age_groups)

    return JsonResponse({
        'labels': list(age_data.keys()),
        'data': list(age_data.values()),
        'colors': ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0']
    })

def get_gender_distribution(request):
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')

    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()

    queryset = Last24Hour.objects.filter(
        hospital_id__in=SuspiciousHospital.objects.values('hospital_id'),
        hospital_type='P',
        preauth_initiated_date__date__gte=start_date, # Added today filter
        preauth_initiated_date__date__lte=end_date, # Added today filter
    )

    if districts:
        queryset = queryset.filter(district_name__in=districts)

    # Gender processing
    gender_data = queryset.values('gender').annotate(
        count=Count('id')
    ).order_by('-count')

    standardized_data = {'Male': 0, 'Female': 0, 'Unknown': 0}
    gender_mappings = {'M': 'Male', 'MALE': 'Male', 'F': 'Female', 'FEMALE': 'Female'}

    for item in gender_data:
        gender = str(item['gender']).strip().upper() if item['gender'] else 'Unknown'
        mapped_gender = gender_mappings.get(gender, 'Unknown')
        
        # Only count as Unknown if not a recognized value
        if mapped_gender == 'Unknown' and gender in ['MALE', 'FEMALE', 'M', 'F']:
            mapped_gender = 'Male' if gender in ['MALE', 'M'] else 'Female'
            
        standardized_data[mapped_gender] += item['count']

    # Prepare response
    labels, data = [], []
    for gender in ['Male', 'Female', 'Unknown']:
        if standardized_data.get(gender, 0) > 0:
            labels.append(gender)
            data.append(standardized_data[gender])

    return JsonResponse({
        'labels': labels,
        'data': data,
        'colors': ['#36A2EB', '#FF6384', '#CCCCCC'][:len(labels)]
    })

SHAPEFILE_DISTRICT_MAPPING = {
    "sheohar": 1,
    "sitamarhi": 2,
    "madhubani": 3,
    "supaul": 4,
    "araria": 5,
    "purnia": 6,
    "katihar": 7,
    "madhepura": 8,
    "saharsa": 9,
    "darbhanga": 10,
    "muzaffarpur": 11,
    "gopalganj": 12,
    "siwan": 13,
    "saran": 14,
    "vaishali": 15,
    "samastipur": 16,
    "begusarai": 17,
    "khagaria": 18,
    "bhagalpur": 19,
    "banka": 20,
    "munger": 21,
    "lakhisarai": 22,
    "sheikhpura": 23,
    "nalanda": 24,
    "patna": 25,
    "bhojpur": 26,
    "kaimur": 27,
    "rohtas": 28,
    "aurangabad": 29,
    "gaya": 30,
    "nawada": 31,
    "jamui": 32,
    "jehanabad": 33,
    "arwal": 34,
    "east champaran": 35,
    "purbi champaran": 35,
    "kishanganj": 36,
    "buxar": 37,
    "west champaran": 38,
}

def get_flagged_claims_geo_counts(request):
    # 1) parse filters
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []
    sd = request.GET.get('start_date')
    ed = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(sd, '%Y-%m-%d').date() if sd else timezone.localdate()
        end_date   = datetime.datetime.strptime(ed, '%Y-%m-%d').date()   if ed else timezone.localdate()
    except ValueError:
        start_date = end_date = timezone.localdate()

    # 2) base queryset
    qs = Last24Hour.objects.filter(
        hospital_id__in=SuspiciousHospital.objects.values('hospital_id'),
        hospital_type='P',
        preauth_initiated_date__date__gte=start_date,
        preauth_initiated_date__date__lte=end_date
    )
    if districts:
        qs = qs.filter(district_name__in=districts)

    # 3) aggregate by district_name
    agg = qs.values('district_name').annotate(count=Count('id'))

    # 4) map back to FID
    result = []
    for row in agg:
        name  = row['district_name']
        cnt   = row['count']
        fid   = SHAPEFILE_DISTRICT_MAPPING.get(name.lower())
        if fid is not None:
            result.append({'fid': fid, 'count': cnt})

    return JsonResponse(result, safe=False)

def download_flagged_claims_excel(request):
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')

    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else date.today()
    except (ValueError, TypeError):
        start_date = date.today()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else date.today()
    except (ValueError, TypeError):
        end_date = date.today()
    
    # 1. Apply same filters as other endpoints
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    # 2. Build queryset with today's filter
    suspicious_hospitals = SuspiciousHospital.objects.values_list('hospital_id', flat=True)
    qs = Last24Hour.objects.filter(
        Q(hospital_id__in=suspicious_hospitals) &
        Q(hospital_type='P') &
        Q(preauth_initiated_date__date__gte=start_date) & 
        Q(preauth_initiated_date__date__lte=end_date)
    )
    
    if districts:
        qs = qs.filter(district_name__in=districts)

    # 3. Prepare data with required fields only
    rows = [{
        'Claim ID': case.registration_id or case.case_id,
        'Patient Name': case.patient_name or f"Patient {case.member_id}",
        'District': case.district_name,
        'Preauth Initiated Date': case.preauth_initiated_date.strftime('%Y-%m-%d'),
        'Preauth Initiated Time': case.preauth_initiated_time,
        'Hospital ID': case.hospital_id,
        'Hospital Name': case.hospital_name,
        'Amount': float(case.claim_initiated_amount) if case.claim_initiated_amount else 0.0,
        'Reason': 'Suspicious hospital',
        'Date': case.preauth_initiated_date.strftime('%Y-%m-%d')
    } for case in qs.only(
        'registration_id', 'case_id', 'patient_name', 'member_id', 'district_name', 'preauth_initiated_date', 'preauth_initiated_time', 'hospital_id', 'hospital_name', 'claim_initiated_amount', 'preauth_initiated_date'
    )]

    # 4. Create DataFrame with defined column order
    columns = ['Claim ID', 'Patient Name', 'District', 'Preauth Initiated Date', 'Preauth Initiated Time', 'Hospital ID', 'Hospital Name',  
              'Amount', 'Reason', 'Date']
    df = pd.DataFrame(rows, columns=columns)

    # 5. Efficient Excel styling
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Flagged Claims')
        workbook = writer.book
        worksheet = writer.sheets['Flagged Claims']
        
        # Style definitions
        red_fill = PatternFill(start_color='FF0000', fill_type='solid')
        white_font = Font(color='FFFFFF')
        
        # Apply styling to entire Reason column
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=9, max_col=9):
            for cell in row:
                cell.fill = red_fill
                cell.font = white_font

    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="flagged_claims_{start_date}_to_{end_date}.xlsx"'
    return response

@require_http_methods(["GET", "POST"])
def download_flagged_claims_report(request):
    startDate = request.POST.get('start_date')
    endDate = request.POST.get('end_date')

    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else date.today()
    except (ValueError, TypeError):
        start_date = date.today()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else date.today()
    except (ValueError, TypeError):
        end_date = date.today()

    # 1) Read parameters & chart images
    district = request.POST.get('district', '')
    districts = district.split(',') if district else []

    # Each value is "data:image/png;base64,XXXXX"
    def strip_prefix(data_url):
        return data_url.split('base64,', 1)[1]

    flagged_b64 = strip_prefix(request.POST.get('flagged_chart', ''))
    age_b64     = strip_prefix(request.POST.get('age_chart', ''))
    gender_b64  = strip_prefix(request.POST.get('gender_chart', ''))
    age_callouts    = request.POST.get('age_callouts', '')
    gender_callouts = request.POST.get('gender_callouts', '')

    # 2) Fetch the FULL flagged-claims data (no pagination)
    suspicious_ids = SuspiciousHospital.objects.values_list('hospital_id', flat=True)
    qs = Last24Hour.objects.filter(
        Q(hospital_id__in=suspicious_ids) &
        Q(hospital_type='P') &
        Q(preauth_initiated_date__date__gte=start_date) &
        Q(preauth_initiated_date__date__lte=end_date)
    )
    if districts:
        qs = qs.filter(district_name__in=districts)

    table_rows = []
    for idx, case in enumerate(qs, start=1):
        table_rows.append({
            'serial_no':     idx,
            'claim_id':      case.registration_id or case.case_id or 'N/A',
            'patient_name':  case.patient_name or f"Patient {case.member_id}",
            'district_name': case.district_name or 'N/A',
            'preauth_initiated_date': case.preauth_initiated_date.strftime('%Y-%m-%d'),
            'preauth_initiated_time': case.preauth_initiated_time,
            'hospital_id': case.hospital_id or 'N/A',
            'hospital_name': case.hospital_name or 'N/A',
            'amount':        case.claim_initiated_amount or 0,
            'reason':        'Suspicious hospital'
        })
    report_districts = sorted({
        row['district_name'] 
        for row in table_rows 
        if row['district_name'] and row['district_name'] != 'N/A'
    })

    # 3) Render HTML via a dedicated template
    context = {
        'logo_url':    request.build_absolute_uri('/static/images/pmjaylogo.png'),
        'title':       'PMJAY FRAUD DETECTION ANALYSIS REPORT',
        'date':        datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'table_rows':  table_rows,
        'report_districts': report_districts,
        'flagged_b64': flagged_b64,
        'age_b64':     age_b64,
        'gender_b64':  gender_b64,
        'age_callouts':    age_callouts,
        'gender_callouts': gender_callouts,
    }
    html_string = render_to_string('flagged_claims_report.html', context)

    # 4) Generate PDF
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    pdf  = html.write_pdf()

    # 5) Return as attachment
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="flagged_claims_report.pdf"'
    return response

def get_high_value_claims(request):
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    if startDate and endDate:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date()
        end_date   = datetime.datetime.strptime(endDate, '%Y-%m-%d').date()
    else:
        today = timezone.localdate()
        start_date = end_date = today
    
    # Base queryset with today's filter
    cases = Last24Hour.objects.filter(
        hospital_type='P',
        preauth_initiated_date__date__gte=start_date,  # Added today filter
        preauth_initiated_date__date__lte=end_date,  # Added today filter
    )
    
    if districts:
        cases = cases.filter(district_name__in=districts)

    # Time thresholds based on today
    yesterday = end_date - timedelta(days=1)
    thirty_days_ago = end_date - timedelta(days=30)

    # Surgical cases filter (≥100,000)
    surgical_cases = cases.filter(
        Q(case_type__iexact='SURGICAL') &
        Q(claim_initiated_amount__gte=100000)
    )

    # Medical cases filter (≥25,000)
    medical_cases = cases.filter(
        Q(case_type__iexact='MEDICAL') &
        Q(claim_initiated_amount__gte=25000)
    )

    # Calculate metrics (now using today-filtered base)
    surgical_total = surgical_cases.aggregate(
        Sum('claim_initiated_amount')
    )['claim_initiated_amount__sum'] or 0
    
    medical_total = medical_cases.aggregate(
        Sum('claim_initiated_amount')
    )['claim_initiated_amount__sum'] or 0

    data = {
        'total_count': surgical_cases.count() + medical_cases.count(),
        'yesterday_count': (
            surgical_cases.filter(preauth_initiated_date__date=yesterday).count() +
            medical_cases.filter(preauth_initiated_date__date=yesterday).count()
        ),
        'last_30_days_count': (
            surgical_cases.filter(preauth_initiated_date__date__gte=thirty_days_ago).count() +
            medical_cases.filter(preauth_initiated_date__date__gte=thirty_days_ago).count()
        ),
        'surgical': {
            'count': surgical_cases.count(),
            'amount': surgical_total,
            'yesterday': surgical_cases.filter(preauth_initiated_date__date=yesterday).count(),
            'last_30_days': surgical_cases.filter(preauth_initiated_date__date__gte=thirty_days_ago).count()
        },
        'medical': {
            'count': medical_cases.count(),
            'amount': medical_total,
            'yesterday': medical_cases.filter(preauth_initiated_date__date=yesterday).count(),
            'last_30_days': medical_cases.filter(preauth_initiated_date__date__gte=thirty_days_ago).count()
        }
    }
    
    return JsonResponse(data)

def get_high_value_claims_details(request):
    case_type = request.GET.get('case_type', 'all').upper()
    district_param = request.GET.get('district', '')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 50))
    
    # 1. Parse start_date / end_date from GET
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()

    # Base query with today's filter
    base_query = Last24Hour.objects.filter(
        hospital_type='P',
        preauth_initiated_date__date__gte=start_date,
        preauth_initiated_date__date__lte=end_date,
    )

    # Case type filters (preserve original logic)
    case_filters = Q()
    if case_type == 'SURGICAL':
        case_filters = Q(case_type__iexact='SURGICAL', claim_initiated_amount__gte=100000)
    elif case_type == 'MEDICAL':
        case_filters = Q(case_type__iexact='MEDICAL', claim_initiated_amount__gte=25000)
    else:
        case_filters = (
            Q(case_type__iexact='SURGICAL', claim_initiated_amount__gte=100000) |
            Q(case_type__iexact='MEDICAL', claim_initiated_amount__gte=25000)
        )
    
    base_query = base_query.filter(case_filters)

    # District filtering
    if district_param:
        districts = district_param.split(',')
        base_query = base_query.filter(district_name__in=districts)

    # Pagination with ordering
    paginator = Paginator(base_query.order_by('-claim_initiated_amount'), page_size)
    page_obj = paginator.get_page(page)

    # Data serialization
    data = []
    for idx, case in enumerate(page_obj.object_list, 1):
        data.append({
            'serial_no': (page_obj.number - 1) * page_size + idx,
            'claim_id': case.registration_id or case.case_id or 'N/A',
            'patient_name': case.patient_name or f"Patient {case.member_id}",
            'district_name': case.district_name or 'N/A',
            'preauth_initiated_date': case.preauth_initiated_date.strftime('%Y-%m-%d'),
            'preauth_initiated_time': case.preauth_initiated_time,
            'hospital_id': case.hospital_id or 'N/A',
            'hospital_name': case.hospital_name or 'N/A',
            'amount': float(case.claim_initiated_amount) if case.claim_initiated_amount else 0.0,
            'case_type': case.case_type.upper() if case.case_type else 'N/A'
        })

    return JsonResponse({
        'data': data,
        'pagination': {
            'total_records': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        }
    })

def get_high_value_claims_by_district(request):
    case_type = request.GET.get('case_type', 'all').upper()
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    # 1. Parse start_date / end_date from GET
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()

    # Base query with today's filter
    base_query = Last24Hour.objects.filter(
        hospital_type='P',
        preauth_initiated_date__date__gte=start_date,  # Added date filter
        preauth_initiated_date__date__lte=end_date,  # Added date filter
    )

    # Apply value thresholds (original logic preserved)
    if case_type == 'SURGICAL':
        base_query = base_query.filter(
            Q(case_type__iexact='SURGICAL') &
            Q(claim_initiated_amount__gte=100000)
        )
    elif case_type == 'MEDICAL':
        base_query = base_query.filter(
            Q(case_type__iexact='MEDICAL') &
            Q(claim_initiated_amount__gte=25000)
        )
    else:  # All
        base_query = base_query.filter(
            Q(case_type__iexact='SURGICAL', claim_initiated_amount__gte=100000) |
            Q(case_type__iexact='MEDICAL', claim_initiated_amount__gte=25000)
        )

    if districts:
        base_query = base_query.filter(district_name__in=districts)

    # Aggregate district data
    district_data = base_query.values('district_name').annotate(
        count=Count('id')
    ).order_by('-count')

    return JsonResponse({
        'districts': [d['district_name'] or 'Unknown' for d in district_data],
        'counts': [d['count'] for d in district_data]
    })

def get_high_value_age_distribution(request):
    case_type = request.GET.get('case_type', 'all').upper()
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []
    # 1. Parse start_date / end_date from GET
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()

    # Base query with today's filter
    base_query = Last24Hour.objects.filter(
        hospital_type='P',
        preauth_initiated_date__date__gte=start_date,  # Added date filter
        preauth_initiated_date__date__lte=end_date  # Added date filter
    )

    # Case type filter (original logic preserved)
    if case_type == 'SURGICAL':
        base_query = base_query.filter(
            Q(case_type__iexact='SURGICAL') &
            Q(claim_initiated_amount__gte=100000)
        )
    elif case_type == 'MEDICAL':
        base_query = base_query.filter(
            Q(case_type__iexact='MEDICAL') &
            Q(claim_initiated_amount__gte=25000)
        )
    else:
        base_query = base_query.filter(
            Q(case_type__iexact='SURGICAL', claim_initiated_amount__gte=100000) |
            Q(case_type__iexact='MEDICAL', claim_initiated_amount__gte=25000)
        )

    if districts:
        base_query = base_query.filter(district_name__in=districts)

    # [Rest of the age grouping logic remains unchanged]
    age_groups = Case(
        When(age_years__lt=20, then=Value('≤20')),
        When(age_years__gte=20, age_years__lt=30, then=Value('21-30')),
        When(age_years__gte=30, age_years__lt=40, then=Value('31-40')),
        When(age_years__gte=40, age_years__lt=50, then=Value('41-50')),
        When(age_years__gte=50, age_years__lt=60, then=Value('51-60')),
        When(age_years__gte=60, then=Value('60+')),
        default=Value('Unknown'),
        output_field=CharField()
    )

    age_data = base_query.annotate(age_group=age_groups).values('age_group') \
        .annotate(count=Count('id')).order_by('age_group')

    categories = ['≤20', '21-30', '31-40', '41-50', '51-60', '60+', 'Unknown']
    colors = ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#FF9F40', '#C9CBCF']
    
    age_dict = {item['age_group']: item['count'] for item in age_data}
    
    return JsonResponse({
        'labels': categories,
        'data': [age_dict.get(cat, 0) for cat in categories],
        'colors': colors
    })

def get_high_value_gender_distribution(request):
    case_type = request.GET.get('case_type', 'all').upper()
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []
    # 1. Parse start_date / end_date from GET
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()

    # Base query with today's filter
    base_query = Last24Hour.objects.filter(
        hospital_type='P',
        preauth_initiated_date__date__gte=start_date,
        preauth_initiated_date__date__lte=end_date
    )

    # Case type filter (original logic preserved)
    if case_type == 'SURGICAL':
        base_query = base_query.filter(
            Q(case_type__iexact='SURGICAL') &
            Q(claim_initiated_amount__gte=100000)
        )
    elif case_type == 'MEDICAL':
        base_query = base_query.filter(
            Q(case_type__iexact='MEDICAL') &
            Q(claim_initiated_amount__gte=25000)
        )
    else:
        base_query = base_query.filter(
            Q(case_type__iexact='SURGICAL', claim_initiated_amount__gte=100000) |
            Q(case_type__iexact='MEDICAL', claim_initiated_amount__gte=25000)
        )

    if districts:
        base_query = base_query.filter(district_name__in=districts)

    # [Rest of the gender grouping logic remains unchanged]
    gender_groups = Case(
        When(gender__iexact='M', then=Value('Male')),
        When(gender__iexact='F', then=Value('Female')),
        When(gender__isnull=False, then=Value('Other')),
        default=Value('Unknown'),
        output_field=CharField()
    )

    gender_data = base_query.annotate(gender_group=gender_groups).values('gender_group') \
        .annotate(count=Count('id')).order_by('gender_group')

    categories = ['Male', 'Female', 'Other', 'Unknown']
    colors = ['#36A2EB', '#FF6384', '#4BC0C0', '#C9CBCF']
    
    gender_dict = {item['gender_group']: item['count'] for item in gender_data}
    
    return JsonResponse({
        'labels': categories,
        'data': [gender_dict.get(cat, 0) for cat in categories],
        'colors': colors
    })

def get_high_value_claims_geo(request):
    case_type = request.GET.get('case_type', 'all').upper()
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    # parse dates
    sd = request.GET.get('start_date')
    ed = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(sd, '%Y-%m-%d').date() if sd else timezone.localdate()
        end_date   = datetime.datetime.strptime(ed, '%Y-%m-%d').date()   if ed else timezone.localdate()
    except ValueError:
        start_date = end_date = timezone.localdate()

    # base queryset
    qs = Last24Hour.objects.filter(
        hospital_type='P',
        preauth_initiated_date__date__gte=start_date,
        preauth_initiated_date__date__lte=end_date
    )
    # thresholds
    if case_type == 'SURGICAL':
        qs = qs.filter(case_type__iexact='SURGICAL', claim_initiated_amount__gte=100000)
    elif case_type == 'MEDICAL':
        qs = qs.filter(case_type__iexact='MEDICAL', claim_initiated_amount__gte=25000)
    else:  # ALL
        qs = qs.filter(
            Q(case_type__iexact='SURGICAL', claim_initiated_amount__gte=100000) |
            Q(case_type__iexact='MEDICAL',  claim_initiated_amount__gte=25000)
        )
    if districts:
        qs = qs.filter(district_name__in=districts)

    # aggregate by district_name
    agg = qs.values('district_name').annotate(count=Count('id'))

    # map to FID
    result = []
    for row in agg:
        fid = SHAPEFILE_DISTRICT_MAPPING.get(row['district_name'].lower())
        if fid is not None:
            result.append({'fid': fid, 'count': row['count']})

    return JsonResponse(result, safe=False)

def get_hospital_bed_cases(request):
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    if startDate and endDate:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date()
        end_date   = datetime.datetime.strptime(endDate, '%Y-%m-%d').date()
    else:
        today = timezone.localdate()
        start_date = end_date = today
    yesterday = end_date - timedelta(days=1)
    thirty_days_ago = end_date - timedelta(days=30)

    # 1. Load bed strengths
    beds = HospitalBeds.objects.values('hospital_id', 'bed_strength')
    bed_strengths = {h['hospital_id']: h['bed_strength'] for h in beds}
    hospital_ids = list(bed_strengths.keys())

    # 2. Helper to fetch admissions per hospital/date
    def admissions_qs(start_date, end_date):
        return Last24Hour.objects.filter(
            Q(hospital_id__in=hospital_ids) &
            (Q(admission_date__date__range=(start_date, end_date)) |
             Q(admission_date__isnull=True, preauth_initiated_date__date__range=(start_date, end_date)))
        )

    # 3. Today: compute per-hospital counts and find violations
    today_adm = admissions_qs(start_date, end_date)
    today_counts = today_adm.values('hospital_id').annotate(count=Count('id'))
    violating_today = [a['hospital_id'] for a in today_counts
                       if a['count'] > bed_strengths.get(a['hospital_id'], 0)]

    # 4. Yesterday: similar
    yest_adm = admissions_qs(yesterday, yesterday)
    yest_counts = yest_adm.values('hospital_id').annotate(count=Count('id'))
    violating_yesterday = [a['hospital_id'] for a in yest_counts
                            if a['count'] > bed_strengths.get(a['hospital_id'], 0)]

    # 5. Last 30 days: annotate per hospital per day, then distinct hospital overflow
    range_adm = admissions_qs(thirty_days_ago, end_date)
    # annotate day and count
    from django.db.models.functions import Coalesce, TruncDate

    annotated = (
        range_adm
        .annotate(ts=Coalesce('admission_date', 'preauth_initiated_date'))
        .annotate(day=TruncDate('ts'))
        .values('hospital_id', 'day')
        .annotate(count=Count('id'))
        .filter(count__gt=0)  # initial filter, we'll apply bed check below
    )
    violating_30_set = set(
        rec['hospital_id']
        for rec in annotated
        if rec['count'] > bed_strengths.get(rec['hospital_id'], 0)
    )

    # 6. Fetch hospital names for today's violations
    names = Last24Hour.objects.filter(
        hospital_id__in=violating_today
    ).values_list('hospital_id', 'hospital_name').distinct()
    hospital_names = {hid: name for hid, name in names}

    # 7. Build detailed list for today's violations
    details_today = []
    for hid in violating_today:
        count = next((a['count'] for a in today_counts if a['hospital_id'] == hid), 0)
        details_today.append({
            'hospital': hospital_names.get(hid, f"Hospital {hid}"),
            'admissions': count,
            'bed_strength': bed_strengths.get(hid, 0)
        })

    # 8. Prepare response
    data = {
        'total': len(violating_today),
        'yesterday': len(violating_yesterday),
        'last_30_days': len(violating_30_set),
        'violations_today': details_today
    }
    return JsonResponse(data)

def get_hospital_bed_details(request):
    district_param = request.GET.get('district', '')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 50))
    districts = district_param.split(',') if district_param else []
    
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()
    
    # 1. Load bed strengths
    beds = HospitalBeds.objects.values('hospital_id', 'bed_strength')
    bed_strengths = {h['hospital_id']: h['bed_strength'] for h in beds}
    
    # 2. Get today's admissions with hospital details
    violations = (
        Last24Hour.objects
        .filter(
            Q(admission_date__date__gte=start_date) &
            Q(admission_date__date__lte=end_date) |
            Q(admission_date__isnull=True, preauth_initiated_date__date__gte=start_date) &
            Q(admission_date__isnull=True, preauth_initiated_date__date__lte=end_date)
        )
        .values('hospital_id', 'hospital_name', 'district_name', 'state_name')
        .annotate(
            admissions=Count('id'),
            last_violation_date=Max('admission_date')
        )
    )
    
    if districts:
        violations = violations.filter(district_name__in=districts)
    
    # 3. Add bed capacity and filter violations
    enhanced_violations = []
    for violation in violations:
        bed_capacity = bed_strengths.get(violation['hospital_id'], 0)
        if violation['admissions'] > bed_capacity:
            enhanced_violations.append({
                **violation,
                'bed_capacity': bed_capacity,
                'excess': violation['admissions'] - bed_capacity
            })
    
    # 4. Paginate in-memory list
    paginator = Paginator(enhanced_violations, page_size)
    page_obj = paginator.get_page(page)
    
    # 5. Serialize data
    data = []
    for idx, violation in enumerate(page_obj.object_list, 1):
        data.append({
            'serial_no': (page_obj.number - 1) * page_size + idx,
            'hospital_id': violation['hospital_id'],
            'hospital_name': violation['hospital_name'],
            'district': violation['district_name'],
            'state': violation['state_name'],
            'bed_capacity': violation['bed_capacity'],
            'admissions': violation['admissions'],
            'excess': violation['excess'],
            'last_violation': violation['last_violation_date'].strftime('%Y-%m-%d') if violation['last_violation_date'] else 'N/A'
        })
    
    return JsonResponse({
        'data': data,
        'pagination': {
            'total_records': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        }
    })

def hospital_violations_by_district(request):
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []
    
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()
    
    result = (
        Last24Hour.objects
        .filter(
            Q(admission_date__date__gte=start_date) &
            Q(admission_date__date__lte=end_date) |
            Q(admission_date__isnull=True, preauth_initiated_date__date__gte=start_date) &
            Q(admission_date__isnull=True, preauth_initiated_date__date__lte=end_date)
        )
        .values('district_name')
        .annotate(violation_count=Count('hospital_id', distinct=True))
        .order_by('-violation_count')
    )
    
    if districts:
        result = result.filter(district_name__in=districts)
    
    return JsonResponse({
        'districts': [item['district_name'] or 'Unknown' for item in result],
        'counts': [item['violation_count'] for item in result]
    })

def get_family_id_cases(request):
    # 1) Parse district filter
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    # 2) Parse date range from GET, default to today
    sd = request.GET.get('start_date')
    ed = request.GET.get('end_date')
    try:
        end_date = datetime.datetime.strptime(ed, '%Y-%m-%d').date() if ed else date.today()
    except (ValueError, TypeError):
        end_date = date.today()
    try:
        start_date = datetime.datetime.strptime(sd, '%Y-%m-%d').date() if sd else end_date
    except (ValueError, TypeError):
        start_date = end_date

    # 3) “Yesterday” and 30‐day window endpoints
    yesterday = end_date - timedelta(days=1)
    thirty_days_ago = end_date - timedelta(days=30)

    # 4) Build base queryset within the window
    base_qs = Last24Hour.objects.filter(
        hospital_type='P',
        preauth_initiated_date__date__gte=start_date,
        preauth_initiated_date__date__lte=end_date
    )
    if districts:
        base_qs = base_qs.filter(district_name__in=districts)

    # 5) Find families with >1 claim *per day* in that window
    suspicious_families = (
        base_qs
        .annotate(day=TruncDate('preauth_initiated_date'))
        .values('family_id', 'day')
        .annotate(count=Count('id'))
        .filter(count__gt=1)
    )

    # 6) “Total” = all cases for those families in the same window
    family_ids = [f['family_id'] for f in suspicious_families]
    cases = base_qs.filter(family_id__in=family_ids)

    # 7) Yesterday’s count: same threshold >1
    yest_qs = Last24Hour.objects.filter(
        hospital_type='P',
        preauth_initiated_date__date=yesterday
    )
    if districts:
        yest_qs = yest_qs.filter(district_name__in=districts)

    yest_families = (
        yest_qs
        .annotate(day=TruncDate('preauth_initiated_date'))
        .values('family_id', 'day')
        .annotate(count=Count('id'))
        .filter(count__gt=1)
        .values_list('family_id', flat=True)
    )
    yesterday_count = yest_qs.filter(family_id__in=yest_families).count()

    # 8) Last 30 days: total cases for those suspicious families
    thr_qs = Last24Hour.objects.filter(
        hospital_type='P',
        preauth_initiated_date__date__gte=thirty_days_ago,
        preauth_initiated_date__date__lte=end_date,
        family_id__in=family_ids
    )
    if districts:
        thr_qs = thr_qs.filter(district_name__in=districts)

    violations_last_30_days = thr_qs.count()

    # 9) Build detailed list for the JSON “violations” array
    violations = []
    for f in suspicious_families:
        hospitals = (
            Last24Hour.objects
            .filter(
                hospital_type='P',
                family_id=f['family_id'],
                preauth_initiated_date__date=f['day']
            )
            .values_list('hospital_name', flat=True)
            .distinct()
        )
        violations.append({
            'family_id': f['family_id'],
            'date':       f['day'].strftime('%Y-%m-%d'),
            'count':      f['count'],
            'hospitals':  list(hospitals)
        })

    # 10) Return the JSON response
    return JsonResponse({
        'total':        cases.count(),
        'yesterday':    yesterday_count,
        'last_30_days': violations_last_30_days,
        'violations':   violations
    })

def get_family_id_cases_details(request):
    district_param = request.GET.get('district', '')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 50))
    districts = district_param.split(',') if district_param else []
    
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()
    
    # Subquery: Get family_ids with more than 2 cases today
    suspicious_families = Last24Hour.objects.annotate(
        day=TruncDate('preauth_initiated_date')
    ).filter(
        day__range=(start_date, end_date)
    ).values('family_id', 'day').annotate(
        count=Count('id')
    ).filter(
        count__gt=1
    ).values('family_id')
    
    # Main query: Cases for those suspicious families today
    cases = Last24Hour.objects.filter(
        hospital_type='P',
        family_id__in=Subquery(suspicious_families),
        preauth_initiated_date__date__gte=start_date,
        preauth_initiated_date__date__lte=end_date
    ).order_by('family_id', 'preauth_initiated_date')
    
    if districts:
        cases = cases.filter(district_name__in=districts)
    
    # Pagination
    paginator = Paginator(cases, page_size)
    page_obj = paginator.get_page(page)
    
    # Serialize data
    data = []
    for idx, case in enumerate(page_obj.object_list, 1):
        data.append({
            'serial_no': (page_obj.number - 1) * page_size + idx,
            'family_id': case.family_id,
            'claim_id': case.registration_id or case.case_id or 'N/A',
            'patient_name': case.patient_name or f"Patient {case.member_id}",
            'district_name': case.district_name or 'N/A',
            'preauth_initiated_date': case.preauth_initiated_date.strftime('%Y-%m-%d'),
            'preauth_initiated_time': case.preauth_initiated_time,
            'hospital_id': case.hospital_id or 'N/A',
            'hospital_name': case.hospital_name or 'N/A',
            'date': case.preauth_initiated_date.date()
        })
    
    return JsonResponse({
        'data': data,
        'pagination': {
            'total_records': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        }
    })


def get_family_violations_by_district(request):
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []
    
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()
    
    # Subquery: Get family_ids with more than 2 cases today
    suspicious_families = Last24Hour.objects.annotate(
        day=TruncDate('preauth_initiated_date')
    ).filter(
        day__range=(start_date, end_date)
    ).values('family_id', 'day').annotate(
        count=Count('id')
    ).filter(
        count__gt=1
    ).values('family_id')
    
    # Main query: Count number of unique families per district
    result = Last24Hour.objects.filter(
        hospital_type='P',
        family_id__in=Subquery(suspicious_families),
        preauth_initiated_date__date__gte=start_date,
        preauth_initiated_date__date__lte=end_date
    )
    
    if districts:
        result = result.filter(district_name__in=districts)
    
    result = result.values('district_name').annotate(
        family_count=Count('family_id', distinct=True)
    ).order_by('-family_count')
    
    return JsonResponse({
        'districts': [item['district_name'] for item in result],
        'counts': [item['family_count'] for item in result]
    })


def get_family_violations_demographics(request, type):
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []
    
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()
    
    # Subquery: Get family_ids with more than 2 cases today
    suspicious_families = Last24Hour.objects.annotate(
        day=TruncDate('preauth_initiated_date')
    ).filter(
        day__range=(start_date, end_date)
    ).values('family_id', 'day').annotate(
        count=Count('id')
    ).filter(
        count__gt=1
    ).values('family_id')
    
    # Base query for demographics
    base_query = Last24Hour.objects.filter(
        hospital_type='P',
        family_id__in=Subquery(suspicious_families),
        preauth_initiated_date__date__gte=start_date,
        preauth_initiated_date__date__lte=end_date
    )
    
    if districts:
        base_query = base_query.filter(district_name__in=districts)
    
    if type == 'age':
        age_groups = Case(
            When(age_years__lt=20, then=Value('≤20')),
            When(age_years__gte=20, age_years__lt=30, then=Value('21-30')),
            When(age_years__gte=30, age_years__lt=40, then=Value('31-40')),
            When(age_years__gte=40, age_years__lt=50, then=Value('41-50')),
            When(age_years__gte=50, age_years__lt=60, then=Value('51-60')),
            When(age_years__gte=60, then=Value('60+')),
            default=Value('Unknown'),
            output_field=CharField()
        )
        
        age_data = base_query.annotate(age_group=age_groups).values('age_group') \
            .annotate(count=Count('id')).order_by('age_group')
        
        categories = ['≤20', '21-30', '31-40', '41-50', '51-60', '60+', 'Unknown']
        colors = ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#FF9F40', '#C9CBCF']
        
        age_dict = {item['age_group']: item['count'] for item in age_data}
        
        data = {
            'labels': categories,
            'data': [age_dict.get(cat, 0) for cat in categories],
            'colors': colors
        }
        
    elif type == 'gender':
        gender_data = base_query.values('gender') \
            .annotate(count=Count('id')).order_by('gender')
        
        categories = ['Male', 'Female', 'Other', 'Unknown']
        colors = ['#36A2EB', '#FF6384', '#4BC0C0', '#C9CBCF']
        
        gender_map = {
            'M': 'Male',
            'F': 'Female',
            'O': 'Other'
        }
        
        gender_dict = {}
        for item in gender_data:
            gender = gender_map.get(item['gender'], 'Unknown')
            gender_dict[gender] = item['count']
        
        data = {
            'labels': categories,
            'data': [gender_dict.get(cat, 0) for cat in categories],
            'colors': colors
        }
    
    return JsonResponse(data)

def get_geo_anomalies(request):
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []
    
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    if startDate and endDate:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date()
        end_date   = datetime.datetime.strptime(endDate, '%Y-%m-%d').date()
    else:
        today = timezone.localdate()
        start_date = end_date = today

    yesterday = end_date - timedelta(days=1)
    thirty_days_ago = end_date - timedelta(days=30)

    # Base queryset - only private hospitals with state mismatches
    anomalies = Last24Hour.objects.filter(
        hospital_type='P',
        state_name__isnull=False,
        hospital_state_name__isnull=False
    ).exclude(state_name=F('hospital_state_name'))

    # Apply district filter (patient's district)
    if districts:
        anomalies = anomalies.filter(district_name__in=districts)

    # Today's count
    today_anomalies = anomalies.filter(
        Q(admission_date__date__gte=start_date) &
        Q(admission_date__date__lte=end_date) | 
        Q(admission_date__isnull=True, preauth_initiated_date__date__gte=start_date) &
        Q(admission_date__isnull=True, preauth_initiated_date__date__lte=end_date)
    )
    
    # Yesterday's count
    yesterday_anomalies = anomalies.filter(
        Q(admission_date__date=yesterday) | 
        Q(admission_date__isnull=True, preauth_initiated_date__date=yesterday)
    )
    
    # Last 30 days count
    last_30_days_anomalies = anomalies.filter(
        Q(admission_date__gte=thirty_days_ago) | 
        Q(admission_date__isnull=True, preauth_initiated_date__gte=thirty_days_ago)
    )

    # Get state mismatch statistics
    state_mismatches = (
        anomalies.values('state_name', 'hospital_state_name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]  # Top 10 mismatch pairs
    )

    data = {
        'total': today_anomalies.count(),
        'yesterday': yesterday_anomalies.count(),
        'last_30_days': last_30_days_anomalies.count(),
        'state_mismatches': list(state_mismatches)
    }

    return JsonResponse(data)

def get_geo_anomalies_details(request):
    district_param = request.GET.get('district', '')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 50))
    districts = district_param.split(',') if district_param else []
    
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()
    
    cases = Last24Hour.objects.filter(
        hospital_type='P',
        state_name__isnull=False,
        hospital_state_name__isnull=False
    ).exclude(state_name=F('hospital_state_name')).filter(
        Q(admission_date__date__gte=start_date) &
        Q(admission_date__date__lte=end_date) | 
        Q(admission_date__isnull=True, preauth_initiated_date__date__gte=start_date) &
        Q(admission_date__isnull=True, preauth_initiated_date__date__lte=end_date)
    ).order_by('state_name', 'hospital_state_name')
    
    if districts:
        cases = cases.filter(district_name__in=districts)

    paginator = Paginator(cases, page_size)
    page_obj = paginator.get_page(page)

    data = []
    for idx, case in enumerate(page_obj.object_list, 1):
        data.append({
            'serial_no': (page_obj.number - 1) * page_size + idx,
            'claim_id': case.registration_id or case.case_id or 'N/A',
            'patient_name': case.patient_name or f"Patient {case.member_id}",
            'district_name': case.district_name or 'N/A',
            'preauth_initiated_date': case.preauth_initiated_date.date() or 'N/A',
            'preauth_initiated_time': case.preauth_initiated_time or 'N/A',
            'hospital_id': case.hospital_id or 'N/A',
            'hospital_name': case.hospital_name or 'N/A',
            'patient_state': case.state_name or 'N/A',
            'hospital_state': case.hospital_state_name or 'N/A',
        })

    return JsonResponse({
        'data': data,
        'pagination': {
            'total_records': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        }
    })

def get_geo_violations_by_state(request):
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []
    
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()
    
    cases = Last24Hour.objects.filter(
        hospital_type='P',
        state_name__isnull=False,
        hospital_state_name__isnull=False
    ).exclude(state_name=F('hospital_state_name')).filter(
        Q(admission_date__date__gte=start_date) &
        Q(admission_date__date__lte=end_date) | 
        Q(admission_date__isnull=True, preauth_initiated_date__date__gte=start_date) &
        Q(admission_date__isnull=True, preauth_initiated_date__date__lte=end_date)
    )
    
    if districts:
        cases = cases.filter(district_name__in=districts)

    result = cases.values('state_name').annotate(count=Count('id')).order_by('-count')
    
    return JsonResponse({
        'states': [item['state_name'] for item in result],
        'counts': [item['count'] for item in result]
    })

def get_geo_violations_demographics(request, type):
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()
    
    base_query = Last24Hour.objects.filter(
        hospital_type='P',
        state_name__isnull=False,
        hospital_state_name__isnull=False
    ).exclude(state_name=F('hospital_state_name')).filter(
        Q(admission_date__date__gte=start_date) &
        Q(admission_date__date__lte=end_date) | 
        Q(admission_date__isnull=True, preauth_initiated_date__date__gte=start_date) &
        Q(admission_date__isnull=True, preauth_initiated_date__date__lte=end_date)
    ).order_by('state_name', 'hospital_state_name')
    
    if districts:
        base_query = base_query.filter(district_name__in=districts)

    if type == 'age':
        age_groups = Case(
            When(age_years__lt=20, then=Value('≤20')),
            When(age_years__gte=20, age_years__lt=30, then=Value('21-30')),
            When(age_years__gte=30, age_years__lt=40, then=Value('31-40')),
            When(age_years__gte=40, age_years__lt=50, then=Value('41-50')),
            When(age_years__gte=50, age_years__lt=60, then=Value('51-60')),
            When(age_years__gte=60, then=Value('60+')),
            default=Value('Unknown'),
            output_field=CharField()
        )
        
        age_data = base_query.annotate(age_group=age_groups).values('age_group') \
            .annotate(count=Count('id')).order_by('age_group')
        
        categories = ['≤20', '21-30', '31-40', '41-50', '51-60', '60+', 'Unknown']
        colors = ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#FF9F40', '#C9CBCF']
        age_dict = {item['age_group']: item['count'] for item in age_data}
        
        data = {
            'labels': categories,
            'data': [age_dict.get(cat, 0) for cat in categories],
            'colors': colors
        }
        
    elif type == 'gender':
        gender_data = base_query.values('gender') \
            .annotate(count=Count('id')).order_by('gender')
        
        categories = ['Male', 'Female', 'Other', 'Unknown']
        colors = ['#36A2EB', '#FF6384', '#4BC0C0', '#C9CBCF']
        gender_map = {'M': 'Male', 'F': 'Female', 'O': 'Other'}
        gender_dict = {gender_map.get(item['gender'], 'Unknown'): item['count'] for item in gender_data}
        
        data = {
            'labels': categories,
            'data': [gender_dict.get(cat, 0) for cat in categories],
            'colors': colors
        }
    
    return JsonResponse(data)

# Module-level cache
df_cache = None
capacity_map = None

def load_dataframes():
    global df_cache, capacity_map
    if df_cache is None:
        # Load hospital capacities
        capacity_qs = SuspiciousHospital.objects.filter(
            number_of_surgeons__isnull=False
        ).values('hospital_id', 'number_of_surgeons')
        
        capacity_map = {
            item['hospital_id']: item['number_of_surgeons'] * 30
            for item in capacity_qs
        }

        # Load last 24h data without datetime combination
        qs = Last24Hour.objects.values(
            'hospital_type',
            'hospital_code',
            'procedure_code',
            'district_name',
            'age_years',
            'patient_name',
            'gender',
            'registration_id',
            'case_id',
            'member_id',
            'hospital_name',
            'preauth_initiated_date',
            'preauth_initiated_time',
            'hospital_id'  # Make sure this exists for capacity checks
        )
        
        df_last = pd.DataFrame.from_records(qs)
        df_cache = df_last
        
    return df_cache, capacity_map


def get_ophthalmology_cases(request):
    # District filter
    districts = request.GET.get('district', '')
    districts = [d.strip() for d in districts.split(',')] if districts else []

    # Date range handling
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        today = timezone.localdate()
        start_date = end_date = today
        
    yesterday = end_date - datetime.timedelta(days=1)
    thirty_days_ago = end_date - datetime.timedelta(days=30)

    # Load data
    df, cap_map = load_dataframes()

    # Convert date strings to datetime objects
    df['preauth_initiated_date'] = pd.to_datetime(df['preauth_initiated_date'])

    # Helper function to create masks
    def period_mask(start_date, end_date):
        mask = (
            (df['hospital_type'] == 'P') &
            (df['procedure_code'].str.contains('SE', na=False)) &
            (df['preauth_initiated_date'].dt.date >= start_date) &
            (df['preauth_initiated_date'].dt.date <= end_date)
        )
        if districts:
            mask &= df['district_name'].isin(districts)
        return mask

    # Create masks for all periods
    mask_total = period_mask(start_date, end_date)
    mask_yest = period_mask(yesterday, yesterday)
    mask_30 = period_mask(thirty_days_ago, end_date)

    # 1. Age <40 counts (unchanged)
    def count_age(mask):
        return int(df.loc[mask & (df['age_years'] < 40)].shape[0])
    
    age_total = count_age(mask_total)
    age_yesterday = count_age(mask_yest)
    age_last_30 = count_age(mask_30)

    # 2. Preauth time validation (simplified)
    def count_preauth(mask):
        # Extract hour directly from time string
        def extract_hour(time_str):
            try:
                if pd.isna(time_str):
                    return None
                return int(time_str.split(':')[0])
            except:
                return None
        
        subset = df.loc[mask].copy()
        subset['hour'] = subset['preauth_initiated_time'].apply(extract_hour)
        
        valid_hours = subset['hour'].notna()
        violations = subset[
            valid_hours & 
            ((subset['hour'] < 8) | (subset['hour'] >= 18))
        ]
        return len(violations)

    preauth_total = count_preauth(mask_total)
    preauth_yesterday = count_preauth(mask_yest)
    preauth_last_30 = count_preauth(mask_30)

    # 3. OT capacity checks (unchanged)
    def compute_ot_exact(df_period):
        ot_total = 0
        for hospital_id, cap in cap_map.items():
            sub = df_period[df_period['hospital_id'] == hospital_id]
            total = len(sub)
            if total > cap:
                ot_total += (total - cap)
        return ot_total
    
    ot_total = compute_ot_exact(df[mask_total])
    ot_yesterday = compute_ot_exact(df[mask_yest])
    ot_last_30 = compute_ot_exact(df[mask_30])

    # Assemble response
    data = {
        'total': ot_total + age_total + preauth_total,
        'age_under_40': {
            'total': age_total,
            'yesterday': age_yesterday,
            'last_30_days': age_last_30
        },
        'ot_cases': {
            'total': ot_total,
            'yesterday': ot_yesterday,
            'last_30_days': ot_last_30,
        },
        'preauth_time': {
            'total': preauth_total,
            'yesterday': preauth_yesterday,
            'last_30_days': preauth_last_30
        }
    }
    
    return JsonResponse(data)

def get_ophthalmology_details(request):
    # 1) Read & normalize GET params
    violation_type = request.GET.get('type', 'all')
    district_param = request.GET.get('district', '').strip()
    districts = [d.strip() for d in district_param.split(',')] if district_param else []

    page      = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 50))

    # 2) Parse start_date / end_date or default to today
    sd = request.GET.get('start_date')
    ed = request.GET.get('end_date')

    try:
        start_date = datetime.datetime.strptime(sd, '%Y-%m-%d').date() if sd else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()

    try:
        end_date = datetime.datetime.strptime(ed, '%Y-%m-%d').date() if ed else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()

    # 3) Load the big DataFrame + capacity map
    df, cap_map = load_dataframes()

    # 4) Ensure preauth date/time are usable
    df['preauth_initiated_date'] = pd.to_datetime(df['preauth_initiated_date'])
    # extract hour from time‐string
    def _extract_hour(ts):
        try:
            return int(ts.split(':')[0])
        except:
            return None
    df['preauth_hour'] = df['preauth_initiated_time'].apply(_extract_hour)

    # 5) Build the “period” mask exactly as in get_ophthalmology_cases
    mask = (
        df['hospital_type'].eq('P') &
        df['procedure_code'].str.contains('SE', na=False) &
        df['preauth_initiated_date'].dt.date.ge(start_date) &
        df['preauth_initiated_date'].dt.date.le(end_date)
    )
    if districts:
        mask &= df['district_name'].isin(districts)

    # 6) Filter & sort by date then time
    df_base = df.loc[mask].copy()
    df_base.sort_values(
        by=['preauth_initiated_date', 'preauth_initiated_time'],
        inplace=True
    )

    # 7) Compute which rows are OT-overflow
    ot_indices = []
    for hosp_id, cap in cap_map.items():
        sub = df_base[df_base['hospital_id'] == hosp_id]
        if len(sub) > cap:
            # everything beyond the first `cap` is flagged
            overflow = sub.iloc[cap:]
            ot_indices += overflow.index.tolist()
    ot_set = set(ot_indices)

    # 8) Build the violation mask per `type`
    if violation_type == 'age':
        vio_mask = df_base['age_years'] < 40
    elif violation_type == 'preauth':
        vio_mask = df_base['preauth_hour'].lt(8) | df_base['preauth_hour'].ge(18)
    elif violation_type == 'ot':
        vio_mask = df_base.index.isin(ot_set)
    elif violation_type == 'multiple':
        m_age     = df_base['age_years'] < 40
        m_preauth = df_base['preauth_hour'].lt(8) | df_base['preauth_hour'].ge(18)
        m_ot      = df_base.index.isin(ot_set)
        vio_mask = (m_age.astype(int) + m_preauth.astype(int) + m_ot.astype(int)) > 1
    else:  # 'all'
        m_age     = df_base['age_years'] < 40
        m_preauth = df_base['preauth_hour'].lt(8) | df_base['preauth_hour'].ge(18)
        m_ot      = df_base.index.isin(ot_set)
        vio_mask = m_age | m_preauth | m_ot

    df_cases = df_base.loc[vio_mask]

    # 9) Paginate
    total_records = len(df_cases)
    start = (page - 1) * page_size
    end   = start + page_size
    page_df = df_cases.iloc[start:end]
    total_pages = (total_records + page_size - 1) // page_size

    # 10) Build response rows
    data = []
    for i, row in enumerate(page_df.itertuples(), start=1):
        serial = start + i
        data.append({
            'serial_no':      serial,
            'claim_id':       row.registration_id or row.case_id or 'N/A',
            'patient_name':   row.patient_name or f"Patient {row.member_id}",
            'hospital_id':    row.hospital_id or 'N/A',
            'hospital_name':  row.hospital_name or 'N/A',
            'district_name':  row.district_name or 'N/A',
            'amount':         getattr(row, 'preauth_initiated_amount', 0) or 0,
            'age':            row.age_years,
            'preauth_time':   row.preauth_initiated_time,
            'age_violation':      (row.age_years < 40) if violation_type in ('age','all','multiple') else None,
            'preauth_violation':  ((row.preauth_hour < 8) or (row.preauth_hour >= 18)) if violation_type in ('preauth','all','multiple') else None,
            'ot_violation':       (row.Index in ot_set) if violation_type in ('ot','all','multiple') else None,
        })

    # 11) Return JSON
    return JsonResponse({
        'data': data,
        'pagination': {
            'total_records': total_records,
            'total_pages':   total_pages,
            'current_page':  page,
            'has_next':      end < total_records,
            'has_previous':  start > 0,
        }
    })

def get_ophthalmology_distribution(request):
    # 1) Params
    violation_type = request.GET.get('type', 'all')
    district_param = request.GET.get('district', '').strip()
    districts = [d.strip() for d in district_param.split(',')] if district_param else []

    # 2) Parse date range or default to today
    sd = request.GET.get('start_date')
    ed = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(sd, '%Y-%m-%d').date() if sd else timezone.localdate()
    except (ValueError, TypeError):
        start_date = timezone.localdate()

    try:
        end_date = datetime.datetime.strptime(ed, '%Y-%m-%d').date() if ed else timezone.localdate()
    except (ValueError, TypeError):
        end_date = timezone.localdate()

    # 3) Load cached DataFrame + capacity map
    df, cap_map = load_dataframes()

    # 4) Prepare date & hour columns
    df['preauth_initiated_date'] = pd.to_datetime(df['preauth_initiated_date'])
    def _extract_hour(ts):
        try:
            return int(ts.split(':')[0])
        except:
            return None
    df['preauth_hour'] = df['preauth_initiated_time'].apply(_extract_hour)

    # 5) Base mask over the requested period
    mask = (
        df['hospital_type'].eq('P') &
        df['procedure_code'].str.contains('SE', na=False) &
        df['preauth_initiated_date'].dt.date.ge(start_date) &
        df['preauth_initiated_date'].dt.date.le(end_date)
    )
    if districts:
        mask &= df['district_name'].isin(districts)

    df_base = df.loc[mask].copy()

    # 6) Compute OT‐overflow indices for the period
    ot_indices = []
    for hosp_id, cap in cap_map.items():
        sub = df_base[df_base['hospital_id'] == hosp_id]
        if len(sub) > cap:
            overflow = sub.iloc[cap:]
            ot_indices.extend(overflow.index.tolist())
    ot_set = set(ot_indices)

    # 7) Build violation mask
    if violation_type == 'age':
        vio_mask = df_base['age_years'] < 40
    elif violation_type == 'preauth':
        vio_mask = df_base['preauth_hour'].lt(8) | df_base['preauth_hour'].ge(18)
    elif violation_type == 'ot':
        vio_mask = df_base.index.isin(ot_set)
    elif violation_type == 'multiple':
        m_age     = df_base['age_years'] < 40
        m_preauth = df_base['preauth_hour'].lt(8) | df_base['preauth_hour'].ge(18)
        m_ot      = df_base.index.isin(ot_set)
        vio_mask = (m_age.astype(int) + m_preauth.astype(int) + m_ot.astype(int)) > 1
    else:  # 'all'
        m_age     = df_base['age_years'] < 40
        m_preauth = df_base['preauth_hour'].lt(8) | df_base['preauth_hour'].ge(18)
        m_ot      = df_base.index.isin(ot_set)
        vio_mask = m_age | m_preauth | m_ot

    df_filtered = df_base.loc[vio_mask]

    # 8) Aggregate by district
    dist_series = df_filtered['district_name'].fillna('Unknown')
    # ensure string dtype
    if dist_series.dtype != object:
        dist_series = dist_series.astype(str)
    counts = dist_series.value_counts()

    districts_out = counts.index.tolist()
    counts_out    = counts.values.tolist()

    # 9) Return JSON
    return JsonResponse({
        'districts': districts_out,
        'counts':    counts_out
    })

def get_ophthalmology_demographics(request, type):
    # 1) Params
    demo_type = type
    violation_type = request.GET.get('violation_type', 'all')
    district_param = request.GET.get('district', '').strip()
    districts = [d.strip() for d in district_param.split(',')] if district_param else []

    # 2) Date range parsing (default to today)
    sd = request.GET.get('start_date')
    ed = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(sd, '%Y-%m-%d').date() if sd else timezone.localdate()
    except (ValueError, TypeError):
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(ed, '%Y-%m-%d').date() if ed else timezone.localdate()
    except (ValueError, TypeError):
        end_date = timezone.localdate()

    # 3) Load cached DF + capacity map
    df, cap_map = load_dataframes()

    # 4) Prepare date/time columns
    df['preauth_initiated_date'] = pd.to_datetime(df['preauth_initiated_date'])
    def _hour(t):
        try:
            return int(t.split(':')[0])
        except:
            return None
    df['preauth_hour'] = df['preauth_initiated_time'].apply(_hour)

    # 5) Base mask over [start_date, end_date]
    mask = (
        df['hospital_type'].eq('P') &
        df['procedure_code'].str.contains('SE', na=False) &
        df['preauth_initiated_date'].dt.date.ge(start_date) &
        df['preauth_initiated_date'].dt.date.le(end_date)
    )
    if districts:
        mask &= df['district_name'].isin(districts)

    df_base = df.loc[mask].copy()

    # 6) Compute OT-overflow set
    ot_indices = []
    for hid, cap in cap_map.items():
        sub = df_base[df_base['hospital_id'] == hid]
        if len(sub) > cap:
            ot_indices += sub.iloc[cap:].index.tolist()
    ot_set = set(ot_indices)

    # 7) Build violation mask
    if violation_type == 'age':
        vio = df_base['age_years'] < 40

    elif violation_type == 'preauth':
        vio = df_base['preauth_hour'].lt(8) | df_base['preauth_hour'].ge(18)

    elif violation_type == 'ot':
        vio = df_base.index.isin(ot_set)

    elif violation_type == 'multiple':
        m1 = df_base['age_years'] < 40
        m2 = df_base['preauth_hour'].lt(8) | df_base['preauth_hour'].ge(18)
        m3 = df_base.index.isin(ot_set)
        vio = (m1.astype(int) + m2.astype(int) + m3.astype(int)) > 1

    else:  # 'all'
        m1 = df_base['age_years'] < 40
        m2 = df_base['preauth_hour'].lt(8) | df_base['preauth_hour'].ge(18)
        m3 = df_base.index.isin(ot_set)
        vio = m1 | m2 | m3

    df_flagged = df_base.loc[vio]

    # 8) Demographic bucketing + counts
    if demo_type == 'age':
        # Define our age buckets
        bins = [0, 20, 30, 40, 50, 60, 200]
        labels = ['≤20','21-30','31-40','41-50','51-60','60+']
        df_flagged['age_group'] = pd.cut(
            df_flagged['age_years'].fillna(-1),
            bins=bins,
            labels=labels,
            right=False
        ).cat.add_categories(['Unknown']).fillna('Unknown')

        counts = df_flagged['age_group'].value_counts().reindex(
            labels + ['Unknown'], fill_value=0
        )

        colors = ['#FF6384','#36A2EB','#FFCE56','#4BC0C0','#9966FF','#FF9F40','#C9CBCF']
        return JsonResponse({
            'labels': labels + ['Unknown'],
            'data':   counts.tolist(),
            'colors': colors
        })

    else:  # gender
        # Map DB codes to labels
        gender_map = {'M':'Male','F':'Female','O':'Other'}
        df_flagged['gender_label'] = df_flagged['gender'].map(gender_map).fillna('Unknown')

        labels = ['Male','Female','Other','Unknown']
        counts = df_flagged['gender_label'].value_counts().reindex(labels, fill_value=0)

        colors = ['#36A2EB','#FF6384','#4BC0C0','#C9CBCF']
        return JsonResponse({
            'labels': labels,
            'data':   counts.tolist(),
            'colors': colors
        })

@require_http_methods(["GET"])
def download_high_value_claims_excel(request):
    # 1) read district filter
    district_param = request.GET.get('district', '')
    districts     = district_param.split(',') if district_param else []
   
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()

    # 2) base queryset for P-type hospitals
    qs = Last24Hour.objects.filter(
        hospital_type='P', 
        preauth_initiated_date__date__gte=start_date,
        preauth_initiated_date__date__lte=end_date
        )
    if districts:
        qs = qs.filter(district_name__in=districts)

    # 3) split into surgical & medical
    surgical_qs = qs.filter(
        case_type__iexact='SURGICAL',
        claim_initiated_amount__gte=100000
    )
    medical_qs = qs.filter(
        case_type__iexact='MEDICAL',
        claim_initiated_amount__gte=25000
    )

    # 4) helper to serialize
    def serialize(qs):
        rows = []
        for idx, c in enumerate(qs, start=1):
            rows.append({
                'S.No':          idx,
                'Claim ID':      c.registration_id or c.case_id or 'N/A',
                'Patient Name':  c.patient_name or f"Patient {c.member_id}",
                'District':      c.district_name or 'N/A',
                'Preauth Initiated Date': c.preauth_initiated_date.strftime('%Y-%m-%d'),
                'Preauth Initiated Time': c.preauth_initiated_time,
                'Hospital ID': c.hospital_id or 'N/A',
                'Hospital Name': c.hospital_name or 'N/A',
                'Amount':        c.claim_initiated_amount or 0,
                'Case Type':     c.case_type.upper() if c.case_type else 'N/A'
            })
        return rows

    df_surgical = pd.DataFrame(serialize(surgical_qs))
    df_medical  = pd.DataFrame(serialize(medical_qs))

    # 5) write to Excel with two sheets
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_surgical.to_excel(writer, index=False, sheet_name='Surgical Reports')
        df_medical.to_excel(writer, index=False, sheet_name='Medical Reports')

        wb   = writer.book
        ws_s = writer.sheets['Surgical Reports']
        ws_m = writer.sheets['Medical Reports']

        # 6) define styles
        red_fill   = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        blue_fill  = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
        white_font = Font(color='FFFFFFFF')

        # 7) style only the “Case Type” cells
        def style_sheet(ws, fill):
            # find the column index of “Case Type” in the header row
            header = next(ws.iter_rows(min_row=1, max_row=1))
            col_idx = next((i+1 for i, cell in enumerate(header) if cell.value == 'Case Type'), None)
            if not col_idx:
                return
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[col_idx-1]
                cell.fill = fill
                cell.font = white_font

        style_sheet(ws_s, red_fill)
        style_sheet(ws_m, blue_fill)

    buffer.seek(0)
    resp = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="high_value_claims_excel_report.xlsx"'
    return resp

@require_POST
@csrf_protect
def download_high_value_claims_report(request):
    # 1) Read inputs
    case_type      = request.POST.get('case_type', 'all').lower()   # 'all','surgical','medical'
    district_param = request.POST.get('district', '')
    districts      = [d for d in district_param.split(',') if d]
    startDate = request.POST.get('start_date')
    endDate = request.POST.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()

    # 2) Helper for charts
    def strip_b64(key):
        val = request.POST.get(key, '')
        return val.split('base64,',1)[1] if 'base64,' in val else ''

    surgical_chart_b64     = strip_b64('surgical_chart')
    surgical_age_chart_b64 = strip_b64('surgical_age_chart')
    surgical_gen_chart_b64 = strip_b64('surgical_gender_chart')
    medical_chart_b64      = strip_b64('medical_chart')
    medical_age_chart_b64  = strip_b64('medical_age_chart')
    medical_gen_chart_b64  = strip_b64('medical_gender_chart')

    # Callouts
    surgical_age_callouts = request.POST.get('surgical_age_callouts','')
    surgical_gen_callouts = request.POST.get('surgical_gender_callouts','')
    medical_age_callouts  = request.POST.get('medical_age_callouts','')
    medical_gen_callouts  = request.POST.get('medical_gender_callouts','')

    # 3) Build querysets based on case_type
    base_qs = Last24Hour.objects.filter(
        hospital_type='P', 
        preauth_initiated_date__date__gte=start_date,
        preauth_initiated_date__date__lte=end_date
        )
    surgical_qs = base_qs.none()
    medical_qs  = base_qs.none()

    if case_type in ['all','surgical']:
        surgical_qs = base_qs.filter(case_type__iexact='SURGICAL', claim_initiated_amount__gte=100000)
    if case_type in ['all','medical']:
        medical_qs  = base_qs.filter(case_type__iexact='MEDICAL',  claim_initiated_amount__gte=25000)

    if districts:
        surgical_qs = surgical_qs.filter(district_name__in=districts)
        medical_qs  = medical_qs.filter(district_name__in=districts)

    # 4) Serialize rows
    surgical_rows = [
        {
            'serial_no':     i+1,
            'claim_id':      c.registration_id or c.case_id or 'N/A',
            'patient_name':  c.patient_name or f"Patient {c.member_id}",
            'district_name': c.district_name or 'N/A',
            'preauth_initiated_date': c.preauth_initiated_date.strftime('%Y-%m-%d'),
            'preauth_initiated_time': c.preauth_initiated_time,
            'hospital_id': c.hospital_id or 'N/A',
            'hospital_name': c.hospital_name or 'N/A',
            'amount':        c.claim_initiated_amount or 0,
            'case_type':     'SURGICAL'
        }
        for i, c in enumerate(surgical_qs)
    ]
    medical_rows = [
        {
            'serial_no':     i+1,
            'claim_id':      c.registration_id or c.case_id or 'N/A',
            'patient_name':  c.patient_name or f"Patient {c.member_id}",
            'district_name': c.district_name or 'N/A',
            'preauth_initiated_date': c.preauth_initiated_date.strftime('%Y-%m-%d'),
            'preauth_initiated_time': c.preauth_initiated_time,
            'hospital_id': c.hospital_id or 'N/A',
            'hospital_name': c.hospital_name or 'N/A',
            'district_name': c.district_name or 'N/A',
            'amount':        c.claim_initiated_amount or 0,
            'case_type':     'MEDICAL'
        }
        for i, c in enumerate(medical_qs)
    ]

    # 5) Compute report_districts as a sorted list
    combined = [r['district_name'] for r in surgical_rows + medical_rows if r['district_name'] and r['district_name'] != 'N/A']
    report_districts = sorted(set(combined))

    # 6) Render
    context = {
        'logo_url':                request.build_absolute_uri('/static/images/pmjaylogo.png'),
        'title':                   'PMJAY FRAUD DETECTION ANALYSIS REPORT',
        'date':                    datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'case_type':               case_type,
        'report_districts':        report_districts,
        'surgical_rows':           surgical_rows,
        'medical_rows':            medical_rows,
        'surgical_chart_b64':      surgical_chart_b64,
        'surgical_age_chart_b64':  surgical_age_chart_b64,
        'surgical_gen_chart_b64':  surgical_gen_chart_b64,
        'medical_chart_b64':       medical_chart_b64,
        'medical_age_chart_b64':   medical_age_chart_b64,
        'medical_gen_chart_b64':   medical_gen_chart_b64,
        'surgical_age_callouts':   surgical_age_callouts,
        'surgical_gen_callouts':   surgical_gen_callouts,
        'medical_age_callouts':    medical_age_callouts,
        'medical_gen_callouts':    medical_gen_callouts,
    }
    html_string = render_to_string('high_value_claims_report.html', context)
    pdf = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="High_Value_Claims_PDF_Report.pdf"'
    return response

@require_http_methods(["GET"])
def download_hospital_bed_cases_excel(request):
    # 1) District filter
    district_param = request.GET.get('district', '')
    districts     = district_param.split(',') if district_param else []

    # 2) date
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()

    # 3) Build bed_strength lookup
    beds = HospitalBeds.objects.values('hospital_id', 'bed_strength')
    bed_strengths = {b['hospital_id']: b['bed_strength'] for b in beds}

    # 4) Query today's admissions + last_violation_date
    qs = (
        Last24Hour.objects
        .filter(
            Q(admission_date__date__gte=start_date) &
            Q(admission_date__date__lte=end_date) |
            Q(admission_date__isnull=True, preauth_initiated_date__date__gte=start_date) &
            Q(admission_date__isnull=True, preauth_initiated_date__date__lte=end_date)
        )
        .values('hospital_id','hospital_name','district_name','state_name')
        .annotate(
            admissions=Count('id'),
            last_violation_date=Max('admission_date')
        )
    )
    if districts:
        qs = qs.filter(district_name__in=districts)

    # 5) Compute “enhanced” where admissions > capacity
    enhanced = []
    for v in qs:
        cap = bed_strengths.get(v['hospital_id'], 0)
        if v['admissions'] > cap:
            enhanced.append({
                'S.No':           len(enhanced) + 1,
                'Hospital ID':    v['hospital_id'],
                'Hospital Name':  v['hospital_name'],
                'District':       v['district_name'],
                'State':          v['state_name'],
                'Bed Capacity':   cap,
                'Admissions':     v['admissions'],
                'Excess':         v['admissions'] - cap,
                'Last Violation': v['last_violation_date'].strftime('%Y-%m-%d')
                                     if v['last_violation_date'] else 'N/A'
            })

    # 6) Build DataFrame
    df = pd.DataFrame(enhanced)

    # 7) Write to Excel with yellow fill on “Excess”
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Hospital Bed Cases')
        ws = writer.sheets['Hospital Bed Cases']

        yellow = PatternFill(start_color='FFFF00',
                             end_color='FFFF00',
                             fill_type='solid')

        # find the column index for “Excess” in row 1
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if 'Excess' in headers:
            col_idx = headers.index('Excess') + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                row[0].fill = yellow

    buffer.seek(0)
    resp = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="hospital_bed_cases.xlsx"'
    return resp

@require_POST
@csrf_protect
def download_hospital_bed_report(request):
    # 1) Read inputs
    district_param = request.POST.get('district','')
    districts = [d for d in district_param.split(',') if d]
    startDate = request.POST.get('start_date')
    endDate = request.POST.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()

    # strip base64
    hc = request.POST.get('hospital_chart','')
    hospital_chart_b64 = hc.split('base64,',1)[1] if 'base64,' in hc else ''

    # 2) Build full violations list (no pagination)
    beds = {b['hospital_id']: b['bed_strength']
            for b in HospitalBeds.objects.values('hospital_id','bed_strength')}

    raw = (
      Last24Hour.objects
      .filter(
        Q(admission_date__date__gte=start_date) &
        Q(admission_date__date__lte=end_date) |
        Q(admission_date__isnull=True, preauth_initiated_date__date__gte=start_date) &
        Q(admission_date__isnull=True, preauth_initiated_date__date__lte=end_date)
      )
      .values('hospital_id','hospital_name','district_name','state_name')
      .annotate(
        admissions=Count('id'),
        last_violation=Max('admission_date')
      )
    )
    if districts:
      raw = raw.filter(district_name__in=districts)

    rows = []
    for i,v in enumerate(raw, start=1):
      cap = beds.get(v['hospital_id'],0)
      if v['admissions'] > cap:
        rows.append({
          'serial_no':      i,
          'hospital_id':    v['hospital_id'],
          'hospital_name':  v['hospital_name'],
          'district':       v['district_name'],
          'state':          v['state_name'],
          'bed_capacity':   cap,
          'admissions':     v['admissions'],
          'excess':         v['admissions']-cap,
          'last_violation': (v['last_violation'].date().isoformat()
                             if v['last_violation'] else 'N/A')
        })

    # 3) report_districts from those rows
    report_districts = sorted({r['district'] for r in rows})

    # 4) Render
    context = {
      'logo_url':         request.build_absolute_uri('/static/images/pmjaylogo.png'),
      'title':            'PMJAY FRAUD DETECTION ANALYSIS REPORT',
      'date':            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
      'report_districts': report_districts,
      'table_rows':       rows,
      'hospital_chart_b64': hospital_chart_b64,
    }
    html = render_to_string('hospital_bed_report.html', context)
    pdf  = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Hospital_Bed_Claims_PDF_Report.pdf"'
    return response

@require_http_methods(["GET"])
def download_family_id_cases_excel(request):
    # 1) district filter
    district_param = request.GET.get('district', '')
    districts     = district_param.split(',') if district_param else []

    # 2) date
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()

    # 3) build the subquery & base queryset
    subq = (
        Last24Hour.objects
        .annotate(day=TruncDate('preauth_initiated_date'))
        .values('family_id', 'day')
        .annotate(count=Count('id'))
        .filter(count__gt=1)
        .values('family_id')
    )
    qs = (
        Last24Hour.objects
        .filter(
            Q(hospital_type='P'),
            Q(family_id__in=Subquery(subq)),
            Q(preauth_initiated_date__date__gte=start_date) &
            Q(preauth_initiated_date__date__lte=end_date)
        )
        .order_by('family_id', 'preauth_initiated_date')
    )
    if districts:
        qs = qs.filter(district_name__in=districts)

    # 4) serialize all rows
    rows = []
    for idx, case in enumerate(qs, start=1):
        rows.append({
            'S.No':          idx,
            'Family ID':     case.family_id,
            'Claim ID':      case.registration_id or case.case_id or 'N/A',
            'Patient Name':  case.patient_name or f"Patient {case.member_id}",
            'District':      case.district_name or 'N/A',
            'Preauth Initiated Date': case.preauth_initiated_date.date().isoformat() or 'N/A',
            'Preauth Initiated Time': case.preauth_initiated_time or 'N/A',
            'Hospital ID': case.hospital_id or 'N/A',
            'Hospital Name': case.hospital_name or 'N/A',
            'Date':          case.preauth_initiated_date.date().isoformat()
        })

    # 5) build DataFrame
    df = pd.DataFrame(rows)

    # 6) write to Excel, then conditionally color “Family ID” cells
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Family ID Cases')
        ws = writer.sheets['Family ID Cases']

        # Only attempt coloring if DataFrame has that column
        if not df.empty and 'Family ID' in df.columns:
            # Map each family_id to a consistent random color
            family_ids = df['Family ID'].unique()
            color_map  = {}
            for fid in family_ids:
                random.seed(str(fid))
                r, g, b = (random.randint(0,255) for _ in range(3))
                color_map[fid] = "{:02X}{:02X}{:02X}".format(r, g, b)

            # Find the column index of “Family ID”
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if 'Family ID' in headers:
                col_idx = headers.index('Family ID') + 1
                yellow_fill = None  # placeholder

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    cell = row[col_idx-1]
                    fid  = cell.value
                    if fid in color_map:
                        fill = PatternFill(
                            start_color=color_map[fid],
                            end_color=color_map[fid],
                            fill_type='solid'
                        )
                        cell.fill = fill

    buffer.seek(0)
    resp = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="family_id_cases.xlsx"'
    return resp

@require_POST
@csrf_protect
def download_family_id_cases_report(request):
    # 1) inputs
    district_param = request.POST.get('district','')
    districts      = [d for d in district_param.split(',') if d]
    startDate = request.POST.get('start_date')
    endDate = request.POST.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else timezone.localdate()
    except ValueError:
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = timezone.localdate()

    # strip base64 helper
    def strip_b64(key):
        val = request.POST.get(key,'')
        return val.split('base64,',1)[1] if 'base64,' in val else ''

    family_chart_b64 = strip_b64('family_chart')
    age_b64          = strip_b64('family_age_chart')
    gender_b64       = strip_b64('family_gender_chart')
    age_callouts     = request.POST.get('age_callouts','')
    gender_callouts  = request.POST.get('gender_callouts','')

    # subquery families with >2 claims today
    freq_families = Last24Hour.objects.annotate(
        day=TruncDate('preauth_initiated_date')
    ).values('family_id','day') \
     .annotate(cnt=Count('id')).filter(cnt__gt=1) \
     .values('family_id')

    qs = Last24Hour.objects.filter(
        Q(hospital_type='P'),
        Q(family_id__in=Subquery(freq_families)),
        Q(preauth_initiated_date__date__gte=start_date) &
        Q(preauth_initiated_date__date__lte=end_date)
    ).order_by('family_id','preauth_initiated_date')

    if districts:
        qs = qs.filter(district_name__in=districts)

    rows = []
    for idx, c in enumerate(qs, start=1):
        rows.append({
            'serial_no':    idx,
            'family_id':    c.family_id,
            'claim_id':     c.registration_id or c.case_id or 'N/A',
            'patient_name': c.patient_name or f"Patient {c.member_id}",
            'district':     c.district_name or 'N/A',
            'preauth_initiated_date':     c.preauth_initiated_date.date().isoformat() or 'N/A',
            'preauth_initiated_time':     c.preauth_initiated_time or 'N/A',
            'hospital_id':     c.hospital_id or 'N/A',
            'hospital_name':     c.hospital_name or 'N/A'
        })

    # 3) report_districts
    report_districts = sorted({r['district'] for r in rows if r['district'] != 'N/A'})

    # 4) render template
    context = {
        'logo_url':            request.build_absolute_uri('/static/images/pmjaylogo.png'),
        'title':               'PMJAY FRAUD DETECTION ANALYSIS REPORT',
        'date':                datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'report_districts':    report_districts,
        'table_rows':          rows,
        'family_chart_b64':    family_chart_b64,
        'age_b64':             age_b64,
        'gender_b64':          gender_b64,
        'age_callouts':        age_callouts,
        'gender_callouts':     gender_callouts,
    }
    html_string = render_to_string('family_id_report.html', context)

    # 5) PDF
    pdf = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="Family_ID_Cases_PDF_Report.pdf"'
    return resp

@require_http_methods(["GET"])
def download_geo_anomalies_excel(request):
    # 1) District filter
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    # 2) date
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else date.today()
    except ValueError:
        start_date = date.today()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = date.today()

    # 3) Query exactly as in get_geo_anomalies_details, but no pagination
    qs = Last24Hour.objects.filter(
        hospital_type='P',
        state_name__isnull=False,
        hospital_state_name__isnull=False
    ).exclude(
        state_name=F('hospital_state_name')
    ).filter(
        Q(admission_date__date__gte=start_date) &
        Q(admission_date__date__lte=end_date) |
        Q(admission_date__isnull=True, preauth_initiated_date__date__gte=start_date) &
        Q(admission_date__isnull=True, preauth_initiated_date__date__lte=end_date)
    ).order_by('state_name','hospital_state_name')

    if districts:
        qs = qs.filter(district_name__in=districts)

    # 4) Serialize all rows
    rows = []
    for idx, c in enumerate(qs, start=1):
        rows.append({
            'S.No':             idx,
            'Claim ID':         c.registration_id or c.case_id or 'N/A',
            'Patient Name':     c.patient_name or f"Patient {c.member_id}",
            'District':         c.district_name or 'N/A',
            'Preauth Initiated Date':         c.preauth_initiated_date.date() or 'N/A',
            'Preauth Initiated Time':         c.preauth_initiated_time or 'N/A',
            'Hospital ID':    c.hospital_id or 'N/A',
            'Hospital Name':    c.hospital_name or 'N/A',
            'Patient State':    c.state_name or 'N/A',
            'Hospital State':   c.hospital_state_name or 'N/A',
        })

    df = pd.DataFrame(rows)

    # 5) Write to Excel, coloring the two state columns per row
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Geo Anomalies')
        ws = writer.sheets['Geo Anomalies']

        # Locate column indexes
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            pat_idx = headers.index('Patient State') + 1
            hos_idx = headers.index('Hospital State') + 1
        except ValueError:
            pat_idx = hos_idx = None

        if pat_idx and hos_idx:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # generate a random color per row
                r, g, b = [random.randint(0,255) for _ in range(3)]
                hexcol = f"{r:02X}{g:02X}{b:02X}"
                fill = PatternFill(start_color=hexcol,
                                   end_color=hexcol,
                                   fill_type='solid')
                # apply to both cells in this row
                row[pat_idx-1].fill = fill
                row[hos_idx-1].fill = fill

    buffer.seek(0)
    resp = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="geographic_anomalies_excel.xlsx"'
    return resp

@require_POST
@csrf_protect
def download_geo_anomalies_pdf_report(request):
    # 1) Read inputs
    district_param = request.POST.get('district','')
    districts      = [d for d in district_param.split(',') if d]

    startDate = request.POST.get('start_date')
    endDate = request.POST.get('end_date')
    try:
        start_date = datetime.datetime.strptime(startDate, '%Y-%m-%d').date() if startDate else date.today()
    except ValueError:
        start_date = date.today()
    try:
        end_date = datetime.datetime.strptime(endDate, '%Y-%m-%d').date() if endDate else timezone.localdate()
    except ValueError:
        end_date = date.today()

    # helper to strip base64
    def strip_b64(key):
        v = request.POST.get(key,'')
        return v.split('base64,',1)[1] if 'base64,' in v else ''

    geo_b64    = strip_b64('geo_chart')
    age_b64    = strip_b64('geo_age_chart')
    gender_b64 = strip_b64('geo_gender_chart')
    age_c      = request.POST.get('geo_age_callouts','')
    gen_c      = request.POST.get('geo_gender_callouts','')

    # 2) Fetch full anomalies (no pagination)
    qs = Last24Hour.objects.filter(
        hospital_type='P',
        state_name__isnull=False,
        hospital_state_name__isnull=False
    ).exclude(state_name=F('hospital_state_name')).filter(
        Q(admission_date__date__gte=start_date) &
        Q(admission_date__date__lte=end_date) |
        Q(admission_date__isnull=True, preauth_initiated_date__date__gte=start_date) &
        Q(admission_date__isnull=True, preauth_initiated_date__date__lte=end_date)
    )
    if districts:
        qs = qs.filter(district_name__in=districts)
    qs = qs.order_by('state_name','hospital_state_name')

    # 3) Build rows
    rows = []
    for i, c in enumerate(qs, start=1):
        rows.append({
            'serial_no':    i,
            'claim_id':     c.registration_id or c.case_id or 'N/A',
            'patient_name': c.patient_name or f"Patient {c.member_id}",
            'district':     c.district_name or 'N/A',
            'preauth_initiated_date':     c.preauth_initiated_date.date() or 'N/A',
            'preauth_initiated_time':     c.preauth_initiated_time or 'N/A',
            'hospital_id':c.hospital_id or 'N/A',
            'hospital_name':c.hospital_name or 'N/A',
            'patient_state':c.state_name or 'N/A',
            'hospital_state':c.hospital_state_name or 'N/A',
        })

    # 4) District line
    report_districts = sorted({r['district'] for r in rows if r['district']!='N/A'})

    # 5) Render PDF
    context = {
      'logo_url':         request.build_absolute_uri('/static/images/pmjaylogo.png'),
      'title':            'PMJAY FRAUD DETECTION ANALYSIS REPORT',
      'date':             datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
      'report_districts': report_districts,
      'table_rows':       rows,
      'geo_chart_b64':    geo_b64,
      'age_chart_b64':    age_b64,
      'gender_chart_b64': gender_b64,
      'age_callouts':     age_c,
      'gender_callouts':  gen_c,
    }
    html_string = render_to_string('geo_anomalies_report.html', context)
    pdf = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()

    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="Geographic_Anomalies_PDF_Report.pdf"'
    return resp

@require_http_methods(["GET"])
def download_ophthalmology_excel(request):
    """
    Download Excel for Ophthalmology (Cataract) violations.
    Supports type filter: all, age, ot, preauth.
    Honors start_date/end_date & district GET params.
    """
    # —1— Read params
    violation_type = request.GET.get('type', 'all')
    district_param = request.GET.get('district', '').strip()
    districts = [d.strip() for d in district_param.split(',')] if district_param else []

    sd = request.GET.get('start_date')
    ed = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(sd, '%Y-%m-%d').date() if sd else timezone.localdate()
    except (ValueError, TypeError):
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(ed, '%Y-%m-%d').date() if ed else timezone.localdate()
    except (ValueError, TypeError):
        end_date = timezone.localdate()

    # —2— Build capacity map
    cap_map = {
        rec['hospital_id']: rec['number_of_surgeons'] * 30
        for rec in SuspiciousHospital.objects
                             .filter(number_of_surgeons__isnull=False)
                             .values('hospital_id', 'number_of_surgeons')
    }

    # —3— Base queryset over the date range
    base_qs = Last24Hour.objects.filter(
        hospital_type='P',
        procedure_code__contains='SE',
        preauth_initiated_date__date__gte=start_date,
        preauth_initiated_date__date__lte=end_date
    )
    if districts:
        base_qs = base_qs.filter(district_name__in=districts)

    # —4— Compute OT-flagged IDs
    flagged_ot_ids = []
    for hid, cap in cap_map.items():
        qs_h = (
            base_qs
            .filter(hospital_id=hid)
            .order_by('preauth_initiated_date', 'preauth_initiated_time')
        )
        total = qs_h.count()
        if total > cap:
            flagged_ot_ids.extend(qs_h.values_list('id', flat=True)[cap:])
    flagged_ot_ids = set(flagged_ot_ids)

    # —5— Helper to choose the right QS for each sheet
    def qs_for(t):
        if t == 'age':
            return base_qs.filter(age_years__lt=40)
        if t == 'ot':
            return base_qs.filter(id__in=flagged_ot_ids)
        if t == 'preauth':
            return base_qs.exclude(
                preauth_initiated_time__regex=r'^(0[8-9]|1[0-7]):'
            )
        # 'all'
        return base_qs.filter(
            Q(age_years__lt=40) |
            Q(id__in=flagged_ot_ids) |
            ~Q(preauth_initiated_time__regex=r'^(0[8-9]|1[0-7]):')
        )

    # —6— Define columns for each sheet
    common = [
        ('S.No',         lambda c,i: i),
        ('Claim ID',     lambda c,i: c.registration_id or c.case_id or 'N/A'),
        ('Patient Name', lambda c,i: c.patient_name or f"Patient {c.member_id}"),
        ('District',     lambda c,i: c.district_name or 'N/A'),
        ('Hospital ID',  lambda c,i: c.hospital_id or 'N/A'),
        ('Hospital Name',lambda c,i: c.hospital_name or 'N/A'),
        ('Amount',       lambda c,i: getattr(c,'preauth_initiated_amount',0) or 0),
    ]

    sheets = {
        'all': {
            'title': 'All cases',
            'fields': common + [
                ('Age<40',       lambda c,i: bool(c.age_years and c.age_years < 40)),
                ('OT Cases',     lambda c,i: c.id in flagged_ot_ids),
                ('Pre-auth Out', lambda c,i: bool(
                    c.preauth_initiated_time and
                    not re.match(r'^(0[8-9]|1[0-7]):', c.preauth_initiated_time)
                ))
            ]
        },
        'age': {
            'title': 'Age <40',
            'fields': common + [
                ('Age<40', lambda c,i: bool(c.age_years and c.age_years < 40))
            ]
        },
        'ot': {
            'title': 'OT Cases',
            'fields': common + [
                ('OT Cases', lambda c,i: c.id in flagged_ot_ids)
            ]
        },
        'preauth': {
            'title': 'Pre-auth Time',
            'fields': common + [
                ('Pre-auth Out', lambda c,i: bool(
                    c.preauth_initiated_time and
                    not re.match(r'^(0[8-9]|1[0-7]):', c.preauth_initiated_time)
                ))
            ]
        }
    }

    # —7— Pick which sheets to emit
    if violation_type in sheets and violation_type != 'all':
        keys = [violation_type]
    else:
        keys = ['all','age','ot','preauth']

    # —8— Build a DataFrame for each sheet
    dfs = {}
    for key in keys:
        rows = []
        qs_iter = qs_for(key).order_by('preauth_initiated_date','preauth_initiated_time')
        for idx, case in enumerate(qs_iter, start=1):
            row = {col: fn(case, idx) for col, fn in sheets[key]['fields']}
            rows.append(row)
        dfs[key] = pd.DataFrame(rows)

    # —9— Write out to Excel with styling
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        fills = {
            'age':      PatternFill('solid', fgColor="FFEB3B"),
            'ot':       PatternFill('solid', fgColor="E91E63"),
            'preauth':  PatternFill('solid', fgColor="009688"),
            'all_high': PatternFill('solid', fgColor="FFCDD2")
        }

        for key in keys:
            title = sheets[key]['title']
            df_sheet = dfs[key]
            df_sheet.to_excel(writer, index=False, sheet_name=title)
            ws = writer.sheets[title]

            # header → column index
            headers = [cell.value for cell in ws[1]]
            idx_map = {h: i+1 for i,h in enumerate(headers)}
            a_i = idx_map.get('Age<40')
            o_i = idx_map.get('OT Cases')
            p_i = idx_map.get('Pre-auth Out')

            for row in ws.iter_rows(min_row=2):
                av = a_i and row[a_i-1].value
                ov = o_i and row[o_i-1].value
                pv = p_i and row[p_i-1].value

                # single‐type fills
                if key in ('age','all') and av and a_i:
                    row[a_i-1].fill = fills['age']
                if key in ('ot','all') and ov and o_i:
                    row[o_i-1].fill = fills['ot']
                if key in ('preauth','all') and pv and p_i:
                    row[p_i-1].fill = fills['preauth']

                # triple overlap only on "all"
                if key=='all' and av and ov and pv:
                    for cell in row:
                        cell.fill = fills['all_high']

    output.seek(0)
    # —10— Build filename with date range
    fname = f"ophthalmology_{violation_type}_{start_date}_to_{end_date}.xlsx"
    resp = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp

@require_POST
@csrf_protect
def download_ophthalmology_pdf_report(request):
    # —1— Read filters
    violation_type = request.POST.get('violation_type', 'all')
    district_param = request.POST.get('district', '').strip()
    districts = [d for d in district_param.split(',') if d]

    # —2— Parse date range (POST) or default to today
    sd = request.POST.get('start_date')
    ed = request.POST.get('end_date')
    try:
        start_date = datetime.datetime.strptime(sd, '%Y-%m-%d').date() if sd else timezone.localdate()
    except (ValueError, TypeError):
        start_date = timezone.localdate()
    try:
        end_date = datetime.datetime.strptime(ed, '%Y-%m-%d').date() if ed else timezone.localdate()
    except (ValueError, TypeError):
        end_date = timezone.localdate()

    # —3— Helper to strip out base64 data-URIs
    def strip_b64(key):
        v = request.POST.get(key, '')
        return v.split('base64,',1)[1] if 'base64,' in v else ''

    # —4— Collect chart images & callouts
    charts = {
        'combined_chart': strip_b64('combined_chart'),
        'age_chart':      strip_b64('age_chart'),
        'ot_chart':       strip_b64('ot_chart'),
        'preauth_chart':  strip_b64('preauth_chart'),
    }
    pies = {}
    callouts = {}
    for section in ['all','age','ot','preauth']:
        pies[f'{section}_age']    = strip_b64(f'{section}_age_chart')
        pies[f'{section}_gender'] = strip_b64(f'{section}_gender_chart')
        callouts[f'{section}_age']    = request.POST.get(f'{section}_age_callouts','')
        callouts[f'{section}_gender'] = request.POST.get(f'{section}_gender_callouts','')

    # —5— Build base queryset over the date range
    base_qs = Last24Hour.objects.filter(
        hospital_type='P',
        procedure_code__contains='SE',
        preauth_initiated_date__date__gte=start_date,
        preauth_initiated_date__date__lte=end_date
    )
    if districts:
        base_qs = base_qs.filter(district_name__in=districts)

    # —6— Compute OT-overflow IDs
    cap_map = {
        rec['hospital_id']: rec['number_of_surgeons'] * 30
        for rec in SuspiciousHospital.objects
                                  .filter(number_of_surgeons__isnull=False)
                                  .values('hospital_id','number_of_surgeons')
    }
    flagged_ot = set()
    for hid, cap in cap_map.items():
        qs_h = (
            base_qs
            .filter(hospital_id=hid)
            .order_by('preauth_initiated_date','preauth_initiated_time')
        )
        total = qs_h.count()
        if total > cap:
            flagged_ot.update(qs_h.values_list('id', flat=True)[cap:])

    # —7— Section‐specific QS
    def section_qs(vtype):
        if vtype == 'age':
            return base_qs.filter(age_years__lt=40)
        if vtype == 'ot':
            return base_qs.filter(id__in=flagged_ot)
        if vtype == 'preauth':
            return base_qs.exclude(
                preauth_initiated_time__regex=r'^(0[8-9]|1[0-7]):'
            )
        # all
        return base_qs.filter(
            Q(age_years__lt=40) |
            Q(id__in=flagged_ot) |
            ~Q(preauth_initiated_time__regex=r'^(0[8-9]|1[0-7]):')
        )

    # —8— Build rows for each section
    rows = {}
    for sec in ['all','age','ot','preauth']:
        qs_iter = section_qs(sec).order_by(
            'preauth_initiated_date','preauth_initiated_time'
        )
        rows[sec] = [
            {
                'serial_no':    i+1,
                'claim_id':     c.registration_id or c.case_id or 'N/A',
                'patient_name': c.patient_name or f"Patient {c.member_id}",
                'district':     c.district_name or 'N/A',
                'hospital_id':  c.hospital_id or 'N/A',
                'hospital_name':c.hospital_name or 'N/A',
                'amount':       getattr(c,'preauth_initiated_amount',0) or 0,
                'age_lt_40':    bool(c.age_years and c.age_years < 40),
                'ot_cases':     c.id in flagged_ot,
                'preauth_time': bool(
                    c.preauth_initiated_time and
                    not re.match(r'^(0[8-9]|1[0-7]):', c.preauth_initiated_time)
                )
            }
            for i, c in enumerate(qs_iter)
        ]

    report_districts = sorted({
        r['district'] for r in rows['all'] if r['district'] != 'N/A'
    })

    # —9— Render template → PDF
    context = {
        'logo_url':         request.build_absolute_uri('/static/images/pmjaylogo.png'),
        'title':            'PMJAY FRAUD DETECTION ANALYSIS REPORT',
        'date':             datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'violation_type':   violation_type,
        'report_districts': report_districts,
        'start_date':       start_date,
        'end_date':         end_date,
        **{f'{sec}_rows': rows[sec] for sec in rows},
        **{f'{k}_b64': v for k, v in charts.items()},
        **{f'{k}_b64': v for k, v in pies.items()},
        **{f'{k}_callouts': v for k, v in callouts.items()},
    }
    html_string = render_to_string('ophthalmology_report.html', context)
    pdf = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()

    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = (
        f'attachment; filename="Ophthalmology_Report_{start_date}_to_{end_date}.pdf"'
    )
    return resp

def high_alert(request):
    sd = request.GET.get('start_date')
    ed = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(sd, '%Y-%m-%d').date() if sd else date.today()
    except (ValueError, TypeError):
        start_date = date.today()
    try:
        end_date = datetime.datetime.strptime(ed, '%Y-%m-%d').date() if ed else date.today()
    except (ValueError, TypeError):
        end_date = date.today()
    district_param = request.GET.get('district', '')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 50))
    districts = district_param.split(',') if district_param else []

    # Get watchlist hospitals
    suspicious_hospitals = SuspiciousHospital.objects.values_list('hospital_id', flat=True)

    # Base query with same filters as get_flagged_claims_details
    base_query = Last24Hour.objects.filter(
        Q(hospital_id__in=suspicious_hospitals) &
        Q(hospital_type='P') &
        (Q(preauth_initiated_date__date__gte=start_date) & Q(preauth_initiated_date__date__lte=end_date) |
         Q(admission_date__date__gte=start_date) & Q(admission_date__date__lte=end_date))
    )

    if districts:
        base_query = base_query.filter(district_name__in=districts)
    annotated_cases = base_query.annotate(
        # Renamed annotations to avoid field conflicts
        is_watchlist=Value(True, output_field=BooleanField()),
        is_high_value=Case(
            When(
                Q(case_type__iexact='SURGICAL', claim_initiated_amount__gte=100000) |
                Q(case_type__iexact='MEDICAL', claim_initiated_amount__gte=25000),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_hospital_bed=Case(
            When(
                Exists(
                    HospitalBeds.objects.filter(
                        hospital_id=OuterRef('hospital_id')
                    ).annotate(
                        admissions=Count(
                            'id',
                            filter=Exists(
                                Last24Hour.objects.filter(
                                    hospital_id=OuterRef('hospital_id'),
                                    admission_date__date__gte=start_date,
                                    admission_date__date__lte=end_date
                                )
                            )
                        )
                    ).filter(admissions__gt=F('bed_strength'))
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_family_case=Case(
            When(
                Exists(
                    Last24Hour.objects.filter(
                        family_id=OuterRef('family_id'),
                        preauth_initiated_date__date__gte=start_date,
                        preauth_initiated_date__date__lte=end_date
                    ).values('family_id').annotate(
                        count=Count('id')
                    ).filter(count__gt=1)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_geo_anomaly=Case(
            When(
                ~Q(state_name=F('hospital_state_name')),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_ophtha_case=Case(
            When(
                Q(procedure_code__contains='SE') & (
                    Q(age_years__lt=40) |
                    Q(preauth_initiated_time__lt=8) |
                    Q(preauth_initiated_time__gte=18)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        )
    )

    # Calculate total flags
    filtered_cases = annotated_cases.annotate(
        total_flags=(
            Cast('is_watchlist', IntegerField()) +
            Cast('is_high_value', IntegerField()) +
            Cast('is_hospital_bed', IntegerField()) +
            Cast('is_family_case', IntegerField()) +
            Cast('is_geo_anomaly', IntegerField()) +
            Cast('is_ophtha_case', IntegerField())
        )
    ).filter(total_flags__gte=2)

    # Pagination
    paginator = Paginator(filtered_cases.order_by('-preauth_initiated_date'), page_size)
    page_obj = paginator.get_page(page)

    # Prepare response data
    data = []
    for idx, case in enumerate(page_obj.object_list, 1):
        data.append({
            'serial_no': (page_obj.number - 1) * page_size + idx,
            'claim_id': case.registration_id or case.case_id or 'N/A',
            'patient_name': case.patient_name or f"Patient {case.member_id}",
            'hospital_id': case.hospital_id,
            'hospital_name': case.hospital_name or 'N/A',
            'district': case.district_name or 'N/A',
            'preauth_initiated_date': case.preauth_initiated_date.strftime('%Y-%m-%d') if case.preauth_initiated_date else 'N/A',
            'preauth_initiated_time': case.preauth_initiated_time or 'N/A',
            'watchlist_hospital': '✓' if case.is_watchlist else '',
            'high_value_claims': '✓' if case.is_high_value else '',
            'hospital_bed_cases': '✓' if case.is_hospital_bed else '',
            'family_id_cases': '✓' if case.is_family_case else '',
            'geographic_anomalies': '✓' if case.is_geo_anomaly else '',
            'ophthalmology_cases': '✓' if case.is_ophtha_case else '',
        })

    return JsonResponse({
        'data': data,
        'pagination': {
            'total_records': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        }
    })

def high_alert_district_distribution(request):
    sd = request.GET.get('start_date')
    ed = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(sd, '%Y-%m-%d').date() if sd else date.today()
    except (ValueError, TypeError):
        start_date = date.today()
    try:
        end_date = datetime.datetime.strptime(ed, '%Y-%m-%d').date() if ed else date.today()
    except (ValueError, TypeError):
        end_date = date.today()
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    # Reuse same filtering as high_alert view
    suspicious_hospitals = SuspiciousHospital.objects.values_list('hospital_id', flat=True)
    
    base_query = Last24Hour.objects.filter(
        Q(hospital_id__in=suspicious_hospitals) &
        Q(hospital_type='P') &
        (Q(preauth_initiated_date__date__gte=start_date) & Q(preauth_initiated_date__date__lte=end_date) | Q(admission_date__date__gte=start_date) & Q(admission_date__date__lte=end_date)))
    if districts:
        base_query = base_query.filter(district_name__in=districts)

    # Apply full high alert criteria annotations
    annotated = base_query.annotate(
        is_watchlist=Value(True, output_field=BooleanField()),
        is_high_value=Case(
            When(
                Q(case_type__iexact='SURGICAL', claim_initiated_amount__gte=100000) |
                Q(case_type__iexact='MEDICAL', claim_initiated_amount__gte=25000),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_hospital_bed=Case(
            When(
                Exists(
                    HospitalBeds.objects.filter(
                        hospital_id=OuterRef('hospital_id')
                    ).annotate(
                        admissions=Count(
                            'id',
                            filter=Exists(
                                Last24Hour.objects.filter(
                                    hospital_id=OuterRef('hospital_id'),
                                    admission_date__date__gte=start_date,
                                    admission_date__date__lte=end_date
                                )
                            )
                        )
                    ).filter(admissions__gt=F('bed_strength'))
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_family_case=Case(
            When(
                Exists(
                    Last24Hour.objects.filter(
                        family_id=OuterRef('family_id'),
                        preauth_initiated_date__date__gte=start_date,
                        preauth_initiated_date__date__lte=end_date
                    ).values('family_id').annotate(
                        count=Count('id')
                    ).filter(count__gt=1)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_geo_anomaly=Case(
            When(
                ~Q(state_name=F('hospital_state_name')),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_ophtha_case=Case(
            When(
                Q(procedure_code__contains='SE') & (
                    Q(age_years__lt=40) |
                    Q(preauth_initiated_time__lt=8) |
                    Q(preauth_initiated_time__gte=18)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        total_flags=(
            Cast('is_watchlist', IntegerField()) +
            Cast('is_high_value', IntegerField()) +
            Cast('is_hospital_bed', IntegerField()) +
            Cast('is_family_case', IntegerField()) +
            Cast('is_geo_anomaly', IntegerField()) +
            Cast('is_ophtha_case', IntegerField())
        )
    ).filter(total_flags__gte=2)

    # Aggregate by district
    result = annotated.values('district_name').annotate(
        case_count=Count('id')
    ).order_by('-case_count')

    return JsonResponse({
        'labels': [d['district_name'] or 'Unknown' for d in result],
        'counts': [d['case_count'] for d in result]
    })

def high_alert_demographics(request, type):
    sd = request.GET.get('start_date')
    ed = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(sd, '%Y-%m-%d').date() if sd else date.today()
    except (ValueError, TypeError):
        start_date = date.today()
    try:
        end_date = datetime.datetime.strptime(ed, '%Y-%m-%d').date() if ed else date.today()
    except (ValueError, TypeError):
        end_date = date.today()
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    # Reuse high alert base query
    suspicious_hospitals = SuspiciousHospital.objects.values_list('hospital_id', flat=True)
    
    base_query = Last24Hour.objects.filter(
        Q(hospital_id__in=suspicious_hospitals) &
        Q(hospital_type='P') &
        (Q(preauth_initiated_date__date__gte=start_date) & Q(preauth_initiated_date__date__lte=end_date) | Q(admission_date__date__gte=start_date) & Q(admission_date__date__lte=end_date))
    )
    
    if districts:
        base_query = base_query.filter(district_name__in=districts)

    # Apply high alert criteria annotations
    annotated = base_query.annotate(
        is_watchlist=Value(True, output_field=BooleanField()),
        is_high_value=Case(
            When(
                Q(case_type__iexact='SURGICAL', claim_initiated_amount__gte=100000) |
                Q(case_type__iexact='MEDICAL', claim_initiated_amount__gte=25000),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_hospital_bed=Case(
            When(
                Exists(
                    HospitalBeds.objects.filter(
                        hospital_id=OuterRef('hospital_id')
                    ).annotate(
                        admissions=Count(
                            'id',
                            filter=Exists(
                                Last24Hour.objects.filter(
                                    hospital_id=OuterRef('hospital_id'),
                                    admission_date__date__gte=start_date,
                                    admission_date__date__lte=end_date
                                )
                            )
                        )
                    ).filter(admissions__gt=F('bed_strength'))
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_family_case=Case(
            When(
                Exists(
                    Last24Hour.objects.filter(
                        family_id=OuterRef('family_id'),
                        preauth_initiated_date__date__gte=start_date,
                        preauth_initiated_date__date__lte=end_date
                    ).values('family_id').annotate(
                        count=Count('id')
                    ).filter(count__gt=1)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_geo_anomaly=Case(
            When(
                ~Q(state_name=F('hospital_state_name')),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_ophtha_case=Case(
            When(
                Q(procedure_code__contains='SE') & (
                    Q(age_years__lt=40) |
                    Q(preauth_initiated_time__lt=8) |
                    Q(preauth_initiated_time__gte=18)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        total_flags=(
            Cast('is_watchlist', IntegerField()) +
            Cast('is_high_value', IntegerField()) +
            Cast('is_hospital_bed', IntegerField()) +
            Cast('is_family_case', IntegerField()) +
            Cast('is_geo_anomaly', IntegerField()) +
            Cast('is_ophtha_case', IntegerField())
        )
    ).filter(total_flags__gte=2)

    if type == 'age':
        age_groups = Case(
            When(age_years__lt=20, then=Value('≤20')),
            When(age_years__range=(20, 29), then=Value('21-30')),
            When(age_years__range=(30, 39), then=Value('31-40')),
            When(age_years__range=(40, 49), then=Value('41-50')),
            When(age_years__range=(50, 59), then=Value('51-60')),
            When(age_years__gte=60, then=Value('60+')),
            default=Value('Unknown'),
            output_field=CharField()
        )
        
        data = annotated.annotate(age_group=age_groups).values('age_group') \
            .annotate(count=Count('id')).order_by('age_group')
        
        categories = ['≤20', '21-30', '31-40', '41-50', '51-60', '60+', 'Unknown']
        colors = ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#FF9F40', '#C9CBCF']
        
        result = {item['age_group']: item['count'] for item in data}
        return JsonResponse({
            'labels': categories,
            'data': [result.get(cat, 0) for cat in categories],
            'colors': colors
        })

    elif type == 'gender':
        gender_data = annotated.values('gender') \
            .annotate(count=Count('id')).order_by('gender')
        
        categories = ['Male', 'Female', 'Other', 'Unknown']
        colors = ['#36A2EB', '#FF6384', '#4BC0C0', '#C9CBCF']
        gender_map = {'M': 'Male', 'F': 'Female', 'O': 'Other'}
        
        result = defaultdict(int)
        for item in gender_data:
            gender = gender_map.get(item['gender'], 'Unknown')
            result[gender] += item['count']
        
        return JsonResponse({
            'labels': categories,
            'data': [result[cat] for cat in categories],
            'colors': colors
        })

def download_high_alerts_excel(request):
    sd = request.GET.get('start_date')
    ed = request.GET.get('end_date')
    try:
        start_date = datetime.datetime.strptime(sd, '%Y-%m-%d').date() if sd else date.today()
    except (ValueError, TypeError):
        start_date = date.today()
    try:
        end_date = datetime.datetime.strptime(ed, '%Y-%m-%d').date() if ed else date.today()
    except (ValueError, TypeError):
        end_date = date.today()
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    # Reuse same filtering logic as high_alert view
    suspicious_hospitals = SuspiciousHospital.objects.values_list('hospital_id', flat=True)
    
    base_query = Last24Hour.objects.filter(
        Q(hospital_id__in=suspicious_hospitals) &
        Q(hospital_type='P') &
        (Q(preauth_initiated_date__date__gte=start_date) & Q(preauth_initiated_date__date__lte=end_date) | Q(admission_date__date__gte=start_date) & Q(admission_date__date__lte=end_date)))
    
    if districts:
        base_query = base_query.filter(district_name__in=districts)

    # Annotate cases with same logic as high_alert view
    annotated_cases = base_query.annotate(
        # Renamed annotations to avoid field conflicts
        is_watchlist=Value(True, output_field=BooleanField()),
        is_high_value=Case(
            When(
                Q(case_type__iexact='SURGICAL', claim_initiated_amount__gte=100000) |
                Q(case_type__iexact='MEDICAL', claim_initiated_amount__gte=25000),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_hospital_bed=Case(
            When(
                Exists(
                    HospitalBeds.objects.filter(
                        hospital_id=OuterRef('hospital_id')
                    ).annotate(
                        admissions=Count(
                            'id',
                            filter=Exists(
                                Last24Hour.objects.filter(
                                    hospital_id=OuterRef('hospital_id'),
                                    admission_date__date__gte=start_date,
                                    admission_date__date__lte=end_date
                                )
                            )
                        )
                    ).filter(admissions__gt=F('bed_strength'))
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_family_case=Case(
            When(
                Exists(
                    Last24Hour.objects.filter(
                        family_id=OuterRef('family_id'),
                        preauth_initiated_date__date__gte=start_date,
                        preauth_initiated_date__date__lte=end_date
                    ).values('family_id').annotate(
                        count=Count('id')
                    ).filter(count__gt=1)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_geo_anomaly=Case(
            When(
                ~Q(state_name=F('hospital_state_name')),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_ophtha_case=Case(
            When(
                Q(procedure_code__contains='SE') & (
                    Q(age_years__lt=40) |
                    Q(preauth_initiated_time__lt=8) |
                    Q(preauth_initiated_time__gte=18)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        )
    )

    # Calculate total flags
    filtered_cases = annotated_cases.annotate(
        total_flags=(
            Cast('is_watchlist', IntegerField()) +
            Cast('is_high_value', IntegerField()) +
            Cast('is_hospital_bed', IntegerField()) +
            Cast('is_family_case', IntegerField()) +
            Cast('is_geo_anomaly', IntegerField()) +
            Cast('is_ophtha_case', IntegerField())
        )
    ).filter(total_flags__gte=2)

    # Prepare data with new columns
    rows = []
    for case in annotated_cases.order_by('-preauth_initiated_date'):
        rows.append({
            'Serial No': '',  # Will be regenerated
            'Claim ID': case.registration_id or case.case_id or 'N/A',
            'Patient Name': case.patient_name or f"Patient {case.member_id}",
            'Hospital ID': case.hospital_id,
            'Hospital Name': case.hospital_name or 'N/A',
            'District': case.district_name or 'N/A',
            'Preauth Initiated Date': case.preauth_initiated_date.strftime('%Y-%m-%d') if case.preauth_initiated_date else 'N/A',
            'Preauth Initiated Time': case.preauth_initiated_time or 'N/A',
            'Watchlist': case.is_watchlist,
            'High Value': case.is_high_value,
            'Bed Cases': case.is_hospital_bed,
            'Family ID': case.is_family_case,
            'Geo Anomaly': case.is_geo_anomaly,
            'Ophthalmology': case.is_ophtha_case
        })

    # Create DataFrame with correct column order
    columns = [
        'Serial No', 'Claim ID', 'Patient Name', 'Hospital ID', 'Hospital Name',
        'District', 'Preauth Initiated Date', 'Preauth Initiated Time',
        'Watchlist', 'High Value', 'Bed Cases', 'Family ID', 'Geo Anomaly', 'Ophthalmology'
    ]
    df = pd.DataFrame(rows, columns=columns)
    df['Serial No'] = df.index + 1  # Continuous numbering

    # Excel styling
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='High Alerts')
        workbook = writer.book
        worksheet = writer.sheets['High Alerts']
        
        # Define color mappings (Excel color codes)
        colors = {
            'Watchlist': '26547D',    # Blue
            'High Value': 'ef436b',    # Red
            'Bed Cases': 'ffce5c',     # Yellow
            'Family ID': '05c793',     # Green
            'Geo Anomaly': '0091b9',   # Dark Blue
            'Ophthalmology': '1abc9c'  # Teal
        }

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply cell coloring
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in colors:
                fill = PatternFill(
                    start_color=colors[col_name],
                    end_color=colors[col_name],
                    fill_type='solid'
                )
                for row in worksheet.iter_rows(
                    min_row=2,
                    max_row=worksheet.max_row,
                    min_col=col_idx,
                    max_col=col_idx
                ):
                    for cell in row:
                        if cell.value:  # Only color if True
                            cell.fill = fill
                            cell.border = thin_border
                            cell.value = ''  # Optional: Clear boolean

    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"high_alerts_{start_date} to_{end_date}_{'_'.join(districts) if districts else 'all'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

from django.views.decorators.csrf import ensure_csrf_cookie
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def download_high_alert_report(request):
    # Get filters and chart data
    district = request.POST.get('district', '')
    districts = district.split(',') if district else []

    sd = request.POST.get('start_date')
    ed = request.POST.get('end_date')
    try:
        start_date = datetime.datetime.strptime(sd, '%Y-%m-%d').date() if sd else date.today()
    except (ValueError, TypeError):
        start_date = date.today()
    try:
        end_date = datetime.datetime.strptime(ed, '%Y-%m-%d').date() if ed else date.today()
    except (ValueError, TypeError):
        end_date = date.today()
    
    # Process chart images
    def strip_prefix(data_url):
        return data_url.split('base64,', 1)[1] if data_url else ''
    
    district_b64 = strip_prefix(request.POST.get('district_chart', ''))
    age_b64 = strip_prefix(request.POST.get('age_chart', ''))
    gender_b64 = strip_prefix(request.POST.get('gender_chart', ''))
    
    # Fetch all high alert cases
    base_query = Last24Hour.objects.filter(
        Q(hospital_id__in=SuspiciousHospital.objects.values_list('hospital_id', flat=True)) &
        Q(hospital_type='P') &
        (Q(preauth_initiated_date__date__gte=start_date) & Q(preauth_initiated_date__date__lte=end_date) | Q(admission_date__date__gte=start_date) & Q(admission_date__date__lte=end_date))
    )
    
    if districts:
        base_query = base_query.filter(district_name__in=districts)
    
    # Annotate and filter
    cases = base_query.annotate(
        # Renamed annotations to avoid field conflicts
        is_watchlist=Value(True, output_field=BooleanField()),
        is_high_value=Case(
            When(
                Q(case_type__iexact='SURGICAL', claim_initiated_amount__gte=100000) |
                Q(case_type__iexact='MEDICAL', claim_initiated_amount__gte=25000),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_hospital_bed=Case(
            When(
                Exists(
                    HospitalBeds.objects.filter(
                        hospital_id=OuterRef('hospital_id')
                    ).annotate(
                        admissions=Count(
                            'id',
                            filter=Exists(
                                Last24Hour.objects.filter(
                                    hospital_id=OuterRef('hospital_id'),
                                    admission_date__date__gte=start_date,
                                    admission_date__date__lte=end_date
                                )
                            )
                        )
                    ).filter(admissions__gt=F('bed_strength'))
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_family_case=Case(
            When(
                Exists(
                    Last24Hour.objects.filter(
                        family_id=OuterRef('family_id'),
                        preauth_initiated_date__date__gte=start_date,
                        preauth_initiated_date__date__lte=end_date
                    ).values('family_id').annotate(
                        count=Count('id')
                    ).filter(count__gt=1)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_geo_anomaly=Case(
            When(
                ~Q(state_name=F('hospital_state_name')),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_ophtha_case=Case(
            When(
                Q(procedure_code__contains='SE') & (
                    Q(age_years__lt=40) |
                    Q(preauth_initiated_time__lt=8) |
                    Q(preauth_initiated_time__gte=18)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        )
    )

    # Calculate total flags
    filtered_cases = cases.annotate(
        total_flags=(
            Cast('is_watchlist', IntegerField()) +
            Cast('is_high_value', IntegerField()) +
            Cast('is_hospital_bed', IntegerField()) +
            Cast('is_family_case', IntegerField()) +
            Cast('is_geo_anomaly', IntegerField()) +
            Cast('is_ophtha_case', IntegerField())
        )
    ).filter(total_flags__gte=2).order_by('-preauth_initiated_date')
    # Prepare table data
    table_rows = []
    for idx, case in enumerate(cases, 1):
        triggered_flags = []
        if case.is_watchlist: triggered_flags.append("Watchlist Hospitals")
        if case.is_high_value: triggered_flags.append("High Value Claims")
        if case.is_hospital_bed: triggered_flags.append("Hospital Bed Violations")
        if case.is_family_case: triggered_flags.append("Family ID Violations")
        if case.is_geo_anomaly: triggered_flags.append("Geo Anomaly")
        if case.is_ophtha_case: triggered_flags.append("Ophthalmology")
        
        table_rows.append({
            'serial_no': idx,
            'claim_id': case.registration_id or case.case_id,
            'patient_name': case.patient_name or f"Patient {case.member_id}",
            'hospital_name': case.hospital_name,
            'district': case.district_name,
            'triggered_flags': triggered_flags,
            'preauth_initiated_date': case.preauth_initiated_date.strftime('%Y-%m-%d'),
            'preauth_initiated_time': case.preauth_initiated_time,
        })
    
    # Render PDF
    context = {
        'logo_url': request.build_absolute_uri('/static/images/pmjaylogo.png'),
        'report_districts': list(set([case['district'] for case in table_rows if case['district']])),
        'table_rows': table_rows,
        'district_b64': district_b64,
        'age_b64': age_b64,
        'gender_b64': gender_b64,
        'date': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    html = render_to_string('high_alert_report.html', context)
    pdf = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="high_alert_report.pdf"'
    return response

@login_required(login_url='login')
def dashboard_view(request):
    # capture the district GET-param so your JS can pick it up
    district_param = request.GET.get('district', '')
    return render(request, 'dashboard.html', {
        'district_param': district_param,
        'active_page': 'dashboard',
    })

@login_required(login_url='login')
def high_alert_view(request):
    district_param = request.GET.get('district', '')
    return render(request, 'high_alert.html', {
        'district_param': district_param,
        'active_page': 'high_alert',
})
